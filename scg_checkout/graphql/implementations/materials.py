import base64
import logging
import os
from datetime import datetime

import pandas as pd
import petl as etl
from dateutil.relativedelta import relativedelta
from django.core.exceptions import ImproperlyConfigured, ValidationError
from django.db import transaction
from django.db.models import Q, Subquery
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from saleor.csv import FileTypes
from saleor.csv.utils.export import create_file_with_headers
from sap_master_data import models as master_models
from sap_migration import models as migration_models
from sap_migration.models import AlternateMaterialOSMappingFileLog
from scg_checkout.error_codes import ContractCheckoutErrorCode
from scg_checkout.graphql.enums import AlternativeMaterial, AlternativeMaterialTypes, AltMaterialType, \
    DistributionChannelType
from scg_checkout.models import AlternatedMaterial, AlternativeMaterialLastUpdateBy

ALT_MAT_LOG_CHANGE_MAX_NO_OF_MONTHS = 3


@transaction.atomic
def import_alternative_material(file_data, user):
    """
    Import alternate material by upload file excel
    All alternate material is Material Variant Master
    @param file_data:
    @param user:
    @return:
    """
    if file_data is None:
        raise ValidationError(
            {
                "file": ValidationError(
                    "No files have been uploaded.",
                    code=ContractCheckoutErrorCode.INVALID.value,
                )
            }
        )

    # add unique text fragment to the file name to prevent file overriding
    file_name, format = os.path.splitext(file_data._name)
    if file_data.size > AlternativeMaterial.MAX_UPLOAD_KB.value:
        raise ValidationError(
            {
                "file_size": ValidationError(
                    "File size is not over 10MB.",
                    code=ContractCheckoutErrorCode.INVALID.value,
                )
            }
        )
    if format != ".xlsx":
        raise ValidationError(
            {
                "file_type": ValidationError(
                    "Only support .xlsx file type. Please try again",
                    code=ContractCheckoutErrorCode.INVALID.value,
                )
            }
        )

    try:
        # Remove all exist items
        migration_models.AlternateMaterial.objects.all().delete()

        expected_order = ["Sale Org.", "Sold to code", "Material Input", "Alternated Material", "Dia", "Type",
                          "Priority"]

        df = pd.read_excel(file_data, nrows=1)

        if df.columns.size < AlternativeMaterial.SHEET_COLUMN_SIZE.value:
            raise ValidationError(
                {
                    "-": ValidationError(
                        f"Invalid format",
                        code=ContractCheckoutErrorCode.INVALID.value,
                    )
                }
            )

        actual_order = df.columns.tolist()

        # Validate if the actual order matches the expected order
        if actual_order != expected_order:
            raise ValidationError(
                {
                    "1": ValidationError(
                        f"Invalid column headers. Expected order is {expected_order}",
                        code=ContractCheckoutErrorCode.INVALID.value,
                    )
                }
            )

        df = pd.read_excel(file_data, dtype={
            0: "string",
            1: "string",
            2: "string",
            3: "string",
            4: "string",
            5: "string",
        }, keep_default_na=False)

        sheet_lines = df.values.tolist()

        # Get list material code and sold to code, and validate blank line
        (
            dict_db_materials_by_material,
            dict_db_materials_by_grade_gram,
            dict_db_organizations,
            dict_db_sold_tos,
            validation_errors,
        ) = get_data_in_db(sheet_lines)

        now = datetime.now()
        alternative_materials = []
        alternative_materials_os = []
        dict_alternative_materials = {}
        handled_lines = []
        for index, line in enumerate(sheet_lines):
            if line not in handled_lines:
                material_input_code = line[2].strip()
                alt_material_code = line[3].strip()
                organization_code = line[0].strip()
                sold_to_code = line[1].strip()
                if organization_code != "":
                    organization_code = organization_code.zfill(AlternativeMaterial.SALE_ORG_LENGTH.value)
                if sold_to_code != "":
                    sold_to_code = sold_to_code.zfill(AlternativeMaterial.SOLD_TO_CODE_LENGTH.value)
                material_type = None
                if line[5].strip() == AlternativeMaterialTypes.MATERIAL.value:
                    material_type = line[5].strip()

                errors = validate_sheet_line(
                    index,
                    line,
                    dict_db_materials_by_material,
                    dict_db_materials_by_grade_gram,
                    dict_db_organizations,
                    dict_db_sold_tos
                )
                if len(errors):
                    validation_errors += errors
                    continue

                material_input = dict_db_materials_by_material.get(material_input_code)
                alt_material = dict_db_materials_by_material.get(alt_material_code)
                if material_type != AlternativeMaterialTypes.MATERIAL.value:
                    material_input = dict_db_materials_by_grade_gram.get(material_input_code)
                    alt_material = dict_db_materials_by_grade_gram.get(alt_material_code)

                organization = dict_db_organizations.get(organization_code)
                sold_to = dict_db_sold_tos.get(sold_to_code)

                # Check alternative material is created or not
                alternative_material = dict_alternative_materials.get(
                    f"{organization_code}_{sold_to_code}_{material_input_code}")
                if not alternative_material:
                    alternative_material = migration_models.AlternateMaterial(
                        sales_organization=organization,
                        sold_to=sold_to,
                        material_own=material_input,
                        type=material_type,
                        created_by=user,
                        updated_by=user,
                        created_at=now,
                        updated_at=now,
                    )
                    # Add to created alternative material
                    alternative_materials.append(alternative_material)
                    dict_alternative_materials[
                        f"{organization_code}_{sold_to_code}_{material_input_code}"] = alternative_material

                diameter = None if material_type == AlternativeMaterialTypes.MATERIAL.value \
                    else line[4].strip().zfill(3)

                alternate_material_os = migration_models.AlternateMaterialOs(
                    alternate_material=alternative_material,
                    material_os=alt_material,
                    diameter=diameter,
                    priority=line[6],
                )
                alternative_materials_os.append(alternate_material_os)

                # Save handled line
                handled_lines.append(line)

        if len(validation_errors):
            validation_errors = sorted(validation_errors, key=lambda x: x["line"])
            raise_validation_errors(validation_errors)

        migration_models.AlternateMaterial.objects.bulk_create(alternative_materials)
        migration_models.AlternateMaterialOs.objects.bulk_create(alternative_materials_os)
        # NOTE: Below code upload to S3 bucket based on which AWS environment is pointed for Local testing.
        file_log_instance = AlternateMaterialOSMappingFileLog.objects.create(
            file_name=file_name,
            file_path=file_data,
            uploaded_by=user,
        )

        last_update_by, created = AlternativeMaterialLastUpdateBy.objects.get_or_create(
            defaults={'updated_by': user}
        )
        if not created:
            last_update_by.updated_by = user
            last_update_by.save()

        logging.info(
            f"[Alternate Material Master Mapping Upload] "
            f"Processed Alt Material Master mappings successfully and file log status id: '{file_log_instance.id}' "
        )
        return True
    except ValidationError as validation_error:
        transaction.set_rollback(True)
        raise validation_error
    except Exception as e:
        transaction.set_rollback(True)
        raise ImproperlyConfigured(e)


def validate_blank_cell(index, line):
    """
    return list errors when line has blank cell
    """
    errors = []
    if line[0].strip() == "":
        errors.append(
            {
                "line": index + 2,
                "message": "Sale Org. cannot be blank",
            }
        )
    if line[1].strip() == "":
        errors.append(
            {
                "line": index + 2,
                "message": "Sold to code cannot be blank",
            }
        )
    elif len(line[1].strip()) > AlternativeMaterial.SOLD_TO_CODE_LENGTH.value or not line[1].strip().isnumeric():
        errors.append(
            {
                "line": index + 2,
                "message": "Invalid Sold to code",
            }
        )
    if line[2].strip() == "":
        errors.append(
            {
                "line": index + 2,
                "message": "Material Input cannot be blank",
            }
        )
    if line[3].strip() == "":
        errors.append(
            {
                "line": index + 2,
                "message": "Alternated Material cannot be blank",
            }
        )
    if line[5].strip() != AlternativeMaterialTypes.MATERIAL.value:
        if not line[4].strip():
            errors.append(
                {
                    "line": index + 2,
                    "message": "Dia cannot be blank",
                }
            )
        else:
            if len(line[4].strip()) > AlternativeMaterial.DIA_LENGTH.value or not line[4].strip().isnumeric():
                errors.append(
                    {
                        "line": index + 2,
                        "message": "Invalid Dia",
                    }
                )
    else:
        if line[4].strip() != "":
            errors.append(
                {
                    "line": index + 2,
                    "message": "Dia must be blank for Type 'M'",
                }
            )
    if line[5].strip() != "" and line[5].strip() != AlternativeMaterialTypes.MATERIAL.value:
        errors.append(
            {
                "line": index + 2,
                "message": f"Type must be {AlternativeMaterialTypes.MATERIAL.value} or leave blank",
            }
        )
    if isinstance(line[6], str) and line[6].strip() == "":
        errors.append(
            {
                "line": index + 2,
                "message": "Priority cannot be blank",
            }
        )
    elif not isinstance(line[6], int) or (isinstance(line[6], int) and (line[6] < 1 or line[6] > 5)):
        errors.append(
            {
                "line": index + 2,
                "message": "Priority must be between 1 and 5",
            }
        )
    return errors


def validate_sheet_line(
        index,
        line,
        dict_db_materials_by_material,
        dict_db_materials_by_grade_gram,
        dict_db_organizations,
        dict_db_sold_tos
):
    """
    return list errors when data don't exist in DB
    """
    errors = []
    material_own_code = line[2].strip()
    material_os_code = line[3].strip()
    organization_code = line[0].strip()
    sold_to_code = line[1].strip()
    material_type = line[5].strip()

    if material_type == AlternativeMaterialTypes.MATERIAL.value:
        if material_own_code != "" and dict_db_materials_by_material.get(material_own_code) is None:
            errors.append(
                {
                    "line": index + 2,
                    "message": "Invalid Material Input",
                }
            )

        if material_os_code != "" and dict_db_materials_by_material.get(material_os_code) is None:
            errors.append(
                {
                    "line": index + 2,
                    "message": "Invalid Alternated Material",
                }
            )
    else:
        if material_own_code != "" and dict_db_materials_by_grade_gram.get(material_own_code) is None:
            errors.append(
                {
                    "line": index + 2,
                    "message": "Invalid Material Input",
                }
            )

        if material_os_code != "" and dict_db_materials_by_grade_gram.get(material_os_code) is None:
            errors.append(
                {
                    "line": index + 2,
                    "message": "Invalid Alternated Material",
                }
            )

    if organization_code != "" and dict_db_organizations.get(organization_code.zfill(AlternativeMaterial.SALE_ORG_LENGTH.value)) is None:
        errors.append(
            {
                "line": index + 2,
                "message": "Invalid Sale Org.",
            }
        )

    if sold_to_code != "" and dict_db_sold_tos.get(sold_to_code.zfill(AlternativeMaterial.SOLD_TO_CODE_LENGTH.value)) is None:
        errors.append(
            {
                "line": index + 2,
                "message": "Invalid Sold to code",
            }
        )
    return errors


def get_data_in_db(sheet_lines):
    """
    Get data from DB to handle line of sheet
    Material or grade gram is material variant code
    """
    material_codes = []
    grade_gram_codes = []
    organization_codes = []
    sold_to_codes = []
    validation_errors = []
    check_duplicated_row = []
    max_priority_count_per_material_input = {}
    for index, line in enumerate(sheet_lines):
        # Check duplicated record by Sale Org., Sold to code, Material - Own, Priority
        organization_code = line[0].strip()
        if organization_code != "":
            organization_code = organization_code.zfill(AlternativeMaterial.SALE_ORG_LENGTH.value)
        sold_to_code = line[1].strip()
        if sold_to_code != "":
            sold_to_code = sold_to_code.zfill(AlternativeMaterial.SOLD_TO_CODE_LENGTH.value)
        unique_line = [organization_code, sold_to_code, line[2].strip(), line[6]]
        if unique_line in check_duplicated_row:
            validation_errors += [
                {
                    "line": index + 2,
                    "message": "Duplicated record",
                }
            ]
            continue
        check_duplicated_row.append(unique_line)

        # Validate blank cell of row
        errors = validate_blank_cell(index, line)
        if len(errors):
            validation_errors += errors

        material_type = line[5].strip()
        if material_type == AlternativeMaterialTypes.MATERIAL.value:
            material_own_code = line[2].strip()
            if material_own_code != "" and material_own_code not in material_codes:
                material_codes.append(material_own_code)

            material_os_code = line[3].strip()
            if material_os_code != "" and material_os_code not in material_codes:
                material_codes.append(material_os_code)
        else:
            grade_gram_own_code = line[2].strip()
            if grade_gram_own_code != "" and grade_gram_own_code not in grade_gram_codes:
                grade_gram_codes.append(grade_gram_own_code)

            grade_gram_os_code = line[3].strip()
            if grade_gram_os_code != "" and grade_gram_os_code not in grade_gram_codes:
                grade_gram_codes.append(grade_gram_os_code)

        if organization_code != "" and organization_code not in organization_codes:
            organization_codes.append(organization_code)

        if sold_to_code != "" and sold_to_code not in sold_to_codes:
            sold_to_codes.append(sold_to_code)

        key = f"{organization_code}:_: {sold_to_code}:_: {line[2].strip()}"
        if not max_priority_count_per_material_input.get(key):
            max_priority_count_per_material_input[key] = 1
        else:
            max_priority_count_per_material_input[key] += 1
            if max_priority_count_per_material_input.get(key) > 5:
                validation_errors += [
                    {
                        "line": index + 2,
                        "message": "Alternated Material exceed maximum of 5 priorities",
                    }
                ]
                continue

    # Get Materials by code
    qs_materials_by_material = master_models.MaterialMaster.objects.filter(
        material_type__in=AltMaterialType.MATERIAL.value,
        material_code__in=Subquery(
            master_models.MaterialSaleMaster.objects.filter(
                distribution_channel_code__in=Subquery(
                    master_models.DistributionChannelMaster.objects.filter(
                        type=DistributionChannelType.DOMESTIC.value).values(
                        'code')),
                material_code__in=material_codes).distinct(
                "material_code").values(
                "material_code")
        )).exclude(delete_flag="X")

    dict_db_materials_by_material = {}
    for m in qs_materials_by_material:
        dict_db_materials_by_material[m.material_code] = m

    # Get Product by grade and gram
    qs_materials_by_grade_gram = master_models.MaterialMaster.objects.filter(
        material_type__in=AltMaterialType.GRADE_GRAM.value,
        material_code__in=Subquery(
            master_models.MaterialSaleMaster.objects.filter(
                distribution_channel_code__in=Subquery(
                    master_models.DistributionChannelMaster.objects.filter(
                        type=DistributionChannelType.DOMESTIC.value).values(
                        'code')), material_code__in=grade_gram_codes).distinct(
                "material_code").values("material_code"))).exclude(delete_flag="X")
    dict_db_materials_by_grade_gram = {}
    for m in qs_materials_by_grade_gram:
        dict_db_materials_by_grade_gram[m.material_code] = m

    qs_organizations = master_models.SalesOrganizationMaster.objects.filter(code__in=organization_codes)
    dict_db_organizations = {}
    for o in qs_organizations:
        dict_db_organizations[o.code] = o

    qs_sold_tos = master_models.SoldToMaster.objects.filter(
        account_group_code__in=["DREP", "Z001"],
        sold_to_code__in=Subquery(
            master_models.SoldToChannelMaster.objects.
            filter(distribution_channel_code__in=["10", "20"],
                   sold_to_code__in=sold_to_codes).distinct(
                "sold_to_code").values("sold_to_code")))

    dict_db_sold_tos = {}
    for s in qs_sold_tos:
        dict_db_sold_tos[s.sold_to_code] = s

    return (
        dict_db_materials_by_material,
        dict_db_materials_by_grade_gram,
        dict_db_organizations,
        dict_db_sold_tos,
        validation_errors,
    )


def raise_validation_errors(validation_errors):
    list_error_messages = {}
    for ve in validation_errors:
        error_line = ve.get("line")
        error_message = ve.get("message")
        if not list_error_messages.get(error_line) or not isinstance(list_error_messages.get(error_line), list):
            list_error_messages[error_line] = [error_message]
        elif error_message not in list_error_messages[error_line]:
            list_error_messages[error_line].append(error_message)

    errors_response = {}
    for line in list_error_messages.keys():
        error_messages = ', '.join(list_error_messages.get(line))
        errors_response[line] = ValidationError(
            error_messages,
            code=ContractCheckoutErrorCode.INVALID.value,
        )

    raise ValidationError(errors_response)


def export_alternative_material_excel_file():
    try:
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        export_file = "Alternated Material.xlsx"
        export_headers = ("Sale Org.", "Sold to code", "Material Input", "Alternated Material", "Dia", "Type",
                          "Priority")
        fields = (
            "alternate_material__sales_organization__code",
            "alternate_material__sold_to__sold_to_code",
            "alternate_material__material_own__material_code",
            "material_os__material_code",
            "diameter",
            "alternate_material__type",
            "priority",
        )

        temp_file = create_file_with_headers(export_headers, delimiter=",", file_type=FileTypes.XLSX)

        count = migration_models.AlternateMaterialOs.objects.count()
        step = 100_000
        for x in range(0, count, step):
            rows = list(migration_models.AlternateMaterialOs.objects.values_list(*fields)[x:x + step])
            etl.io.xlsx.appendxlsx([export_headers] + rows, temp_file.name)

        return export_file, content_type, base64.b64encode(temp_file.read()).decode("utf-8")
    except Exception as e:
        logging.error(e)
        raise ImproperlyConfigured("Internal Server Error.")


CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
EXPORT_FILE = "alternate_log_change_material.xlsx"
DATE_FORMAT = "%d/%m/%Y"
COLUMN_WIDTHS = {
    "H": 20,  # Width for "substitute product" column
    "I": 20,  # Width for "No. of Rolls changed" column
    "J": 20,  # Width for "Amount of change (TON)" column
    "K": 20,  # Width for "plant" column
}


def export_alternative_material_log_excel_file():
    try:
        today = datetime.now()
        date_range = today - relativedelta(months=ALT_MAT_LOG_CHANGE_MAX_NO_OF_MONTHS)
        max_date_range = datetime(date_range.year, date_range.month, 1)

        filtered_rows = AlternatedMaterial.objects.filter(
            Q(order__updated_at__gte=max_date_range) | Q(order__updated_at__isnull=True,
                                                         order__created_at__gte=max_date_range)
        )
        export_headers = (
            "S/O No.", "Item No.", "Sale Organization", "Sold To", "Customer Name", "รหัสสินค้าที่สั่ง", "จำนวนที่สั่ง",
            "รหัสสินค้าทดแทน", "จำนวนม้วนที่เปลี่ยน", "จำนวนที่เปลี่ยน (TON)", "plant",
            "ชื่อไฟล์  (สำหรับ PO Upload)", "PO No.", "Item Request Date", "Create Date", "Create by",
            "Last Update Date",
            "Last Update by"
        )
        fields = (
            "order__so_no",  # index: 0
            "order_line__item_no",  # index: 1
            "order__sales_organization__code",  # index: 2
            "order__sales_organization__name",  # index: 3
            "order__sold_to__sold_to_code",  # index: 4
            "order__sold_to__sold_to_name",  # index: 5
            "old_product__code",  # index: 6
            "order_line__quantity",  # index: 7
            "error_type",  # index: 8
            "quantity_change_of_roll",  # index: 9
            "order_line__plant",  # index: 10
            "order__po_upload_file_log_id__file_name",  # index: 11
            "order__po_no",  # index: 12
            "order_line__request_date",  # index: 13
            "order__created_at",  # index: 14
            "order__created_by_id__first_name",  # index: 15
            "order__updated_at",  # index: 16
            "order__update_by_id__first_name",  # index: 17
            "quantity_change_of_ton"  # index: 18
        )

        temp_file = create_file_with_headers(export_headers, delimiter=",", file_type=FileTypes.XLSX)

        count = filtered_rows.count()
        step = 100_000
        fields_tuple = tuple(fields)

        for x in range(0, count, step):
            rows = list(filtered_rows.values_list(*fields)[x:x + step])
            combined_rows = [{**dict(zip(export_headers, (
                row[0], row[1], f"{row[2]}-{row[3]}", f"{row[4]}-{row[5]}", row[5], row[6], row[7],
                row[8] if row[8] else row[9], row[9], row[18], row[10], row[11], row[12],
                row[13].strftime(DATE_FORMAT) if row[13] else '', row[14].strftime(DATE_FORMAT) if row[14] else '',
                row[15], row[16].strftime(DATE_FORMAT) if row[16] else '', row[17]
            )))} for row in rows]
            excel_rows = []
            # Add an empty row (a list of empty strings) to the excel_rows list
            empty_row = [""] * len(export_headers)
            excel_rows.append(empty_row)

            for y in combined_rows:
                excel_rows.append(y.values())

            etl.io.xlsx.appendxlsx([export_headers], temp_file.name, write_header=True)
            etl.io.xlsx.appendxlsx([list(values) for values in excel_rows], temp_file.name)

            workbook = load_workbook(temp_file.name)
            sheet = workbook.active

            # Merge cells for row1 and row2 in all columns except H to K
            start_col = 8
            end_col = 12
            merge_row_headers(sheet, export_headers, start_col, end_col)

            sheet.merge_cells('H1:K1')
            sheet['H1'].value = "จำนวนที่สั่ง"
            sheet['H1'].alignment = Alignment(horizontal='center', vertical='center')

            merge_columns = {
                "รหัสสินค้าทดแทน": "H2:H2",
                "จำนวนม้วนที่เปลี่ยน": "I2:I2",
                "จำนวนที่เปลี่ยน (TON)": "J2:J2",
                "plant": "K2:K2",
            }

            # Merge cells under the "จำนวนที่สั่ง" column for sub-headers

            merge_sub_headers(sheet, merge_columns)

            for col_num, header in enumerate(export_headers, 1):
                column_letter = get_column_letter(col_num)
                column_width = max(len(str(header)), len(str(sheet.cell(row=2, column=col_num).value))) + 2
                sheet.column_dimensions[column_letter].width = column_width

            workbook.save(temp_file.name)

        with open(temp_file.name, "rb") as exported_file:
            base64_file = base64.b64encode(exported_file.read())

        return EXPORT_FILE, CONTENT_TYPE, base64_file.decode("utf-8")

    except Exception as e:
        logging.error(e)
        raise ImproperlyConfigured("Internal Server Error.")


def merge_row_headers(sheet, export_headers, start_col, end_col):
    num_cols = len(export_headers)
    for col_num in range(1, num_cols + 1):
        col_letter = get_column_letter(col_num)
        if col_num not in range(start_col, end_col):  # Exclude columns H to K
            sheet.merge_cells(f"{col_letter}1:{col_letter}2")
            sheet[f"{col_letter}1"].value = export_headers[col_num - 1]


def merge_sub_headers(sheet, merge_columns):
    for col_header, cell_range in merge_columns.items():
        sheet.merge_cells(cell_range)
        col_start, col_end = cell_range.split(":")
        sheet[col_start].value = col_header


def get_diameter_from_code(full_code):
    """
    Get diameter from material variant full code
    @param full_code:
    @return: string diameter
    """
    try:
        return full_code[14:17]
    except:
        return ""
