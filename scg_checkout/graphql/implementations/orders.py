import json
import shutil
import time
import uuid
from copy import deepcopy
from math import floor

import boto3
import logging
import base64

from sympy import true
from scg_checkout.graphql.helper import get_sold_to_partner
from scg_checkout.graphql.enums import ContractCheckoutErrorCode

from common.product_group import ProductGroup
from common.weight_calculation import resolve_weight_common
from sap_master_data.models import SoldToMaterialMaster
from sap_migration.implementations.class_mark import class_mark_logic, update_class_mark_to_sap
import scg_checkout.contract_order_update as fn
import openpyxl
from django.utils import timezone
import pytz
from celery.utils.log import get_task_logger
from django.core.exceptions import ImproperlyConfigured, ValidationError
from django.db import transaction, models
from django.db.models import *

from datetime import datetime
import petl as etl
from scg_checkout.graphql.helper import get_name_from_sold_to_partner_address_master,get_sold_to_partner
from common.helpers import parse_json_if_possible, snake_to_camel, net_price_calculation
from common.enum import MulesoftServiceType, EorderingItemStatusEN, EorderingItemStatusTH
from common.mulesoft_api import MulesoftApiRequest
from saleor.csv.utils.export import create_file_with_headers
from saleor.plugins.manager import get_plugins_manager
from sap_migration import models as sap_migration_models
from sap_master_data import models as sap_master_data_modes
from sap_migration.graphql.enums import OrderType, InquiryMethodType
from openpyxl.styles import Alignment
from sap_migration.models import Order
from scg_checkout.graphql.enums import (
    ScgOrderStatus,
    ProductionStatus,
    IPlanOrderItemStatus,
    SapOrderConfirmationStatus,
    ScgOrderStatusSAP,
    DeliveryStatus,
    IPlanTypeOfDelivery, AtpCtpStatus, PendingOrderFieldHeaderColumn, WeightUnitEnum
)
from django.db.models import F, Sum
from openpyxl.styles import Font

from scg_checkout.graphql.implementations.change_order import (
    get_iplan_error_messages
)
from scg_checkout.graphql.implementations.customer_materials import CUSTOMER_MATERIAL_TEMPLATE_RELATIVE_PATH, \
    CUSTOMER_MATERIAL_TEMPLATE_XLSX_FILE_NAME
from scg_checkout.graphql.implementations.dtp_dtr import NOT_PASS_VALUE
from scg_checkout.graphql.implementations.iplan import (
    change_parameter_follow_inquiry_method,
    handle_atp_ctp
)
from scg_checkout.graphql.implementations.sap import get_error_messages_from_sap_response_for_change_order
from scg_checkout.graphql.resolves.orders import call_sap_es26
from scgp_export.graphql.enums import SapEnpoint
from scgp_export.graphql.helper import (
    handle_order_lines_from_es26,
    make_order_text_mapping,
    sync_export_order_from_es26
)
from scgp_export.implementations.edit_order import flag_r5_failed_items
from scgp_export.implementations.iplan import handle_case_being_process_sap
from scgp_export.implementations.orders import (
    mock_confirmed_date,
    get_partner_from_order_for_sending_email,get_sold_to_name_and_address_from_order_for_sending_email,
    make_surname_for_send_order_email,
    _update_reject_reason_for_items, )
from common.iplan.item_level_helpers import get_product_code
from scgp_po_upload.graphql.helpers import html_to_pdf, html_to_pdf_order_confirmation
from scg_checkout.graphql.helper import update_order_status, update_dtr_dtp_to_sap, deepgetattr, is_default_sale_unit_from_contract, is_other_product_group, get_parent_directory, mapping_order_partners, is_materials_product_group_matching, update_order_product_group
from scgp_require_attention_items.graphql.enums import IPlanEndpoint
from scgp_require_attention_items.graphql.helper import (
    add_class_mark_into_order_line,
    update_attention_type_r5,
    scg_round, camel_to_snake, multiple_sort_for_pending_order, set_style
)
from scgp_require_attention_items.graphql.resolvers.require_attention_items import \
    resolve_list_of_sale_order_sap_order_pending

from sap_migration.implementations.dtr_dtp import calculate_dtp_dtr

from common.sap.sap_api import SapApiRequest

import os

logger = logging.getLogger(__name__)

task_logger = get_task_logger(__name__)
ASIA_BANGKOK_TIMEZONE = "Asia/Bangkok"
TIME_FORMAT = "%Y%m%d%H%M"


def validate_order(order_id):
    order = sap_migration_models.Order.objects.get(Q(id=order_id) | Q(so_no=order_id))
    if order.status == ScgOrderStatus.CONFIRMED.value:
        raise ValidationError("Confirmed order can not change status!")
    if not order:
        raise ValidationError(f"Order with id {order_id} not found!")
    return order


def validate_lines_quantity(order_id, contract_material_ids):
    lines = sap_migration_models.OrderLines.all_objects.filter(
        order_id=order_id,
        draft=True
    )
    if contract_material_ids:
        lines = lines.filter(contract_material_id__in=contract_material_ids)
    total_weight_field = models.DecimalField(max_digits=10, decimal_places=3)

    contract_materials = lines.values("contract_material").order_by("contract_material").annotate(sum_quantity=Case(
        When(total_weight__isnull=False, then=Sum(F("weight") * F("quantity"))),
        default=Value(0.0),
        output_field=total_weight_field)).filter(
        sum_quantity__gt=F("contract_material__remaining_quantity")
    )
    contract_material_ids = [
        contract_product.get("contract_material")
        for contract_product in contract_materials
    ]
    if len(contract_material_ids):
        invalid_quantity_line_ids = sap_migration_models.OrderLines.objects.filter(
            order_id=order_id,
            contract_material_id__in=contract_material_ids,
        ).values_list("id", flat=True)
        return invalid_quantity_line_ids
    return []


def combine_two_product_lists(list1, list2):
    combined_list = []
    for product1 in list1:
        for product2 in list2:
            if str(product2.get("contract_material_id")) == str(
                    product1.get("contract_material_id")
            ) and str(product2.get("material_variant_id")) == str(product1.get("material_variant_id")):
                combined_list.append(
                    {
                        "contract_material_id": product2.get("contract_material_id"),
                        "material_variant_id": product2.get("material_variant_id"),
                    }
                )

    return combined_list


@transaction.atomic
def add_products_to_domestic_order(order_id, products):
    try:
        order = sap_migration_models.Order.objects.filter(so_no=order_id).first()
        if len(order_id) < 10:
            order = sap_migration_models.Order.objects.get(id=order_id)
        order_id = order.id
        validate_order(order_id)
        contract_material_ids = []
        material_variant_ids = []
        material_variant_codes = []
        for product in products:
            contract_material_ids.append(product.get("contract_material_id", None))
            if product.get("material_variant_id", ""):
                material_variant_ids.append(product.get("material_variant_id", None))

        sold_to_code = order.sold_to.sold_to_code
        sales_organization = order.sales_organization
        distribution_channel = order.distribution_channel
        channel_master = sap_master_data_modes.SoldToChannelMaster.objects.filter(
            sold_to_code=sold_to_code,
            sales_organization_code=sales_organization.code,
            distribution_channel_code=distribution_channel.code,
        ).first()
        line_objects = {}
        item_no_max = 0
        if order.item_no_latest:
            item_no_max = int(order.item_no_latest)

        for line in sap_migration_models.OrderLines.all_objects.filter(order_id=order_id):
            line_objects[
                str(line.contract_material_id) + "-" + str(line.material_variant_id)] = line
            if int(item_no_max) < int(line.item_no):
                item_no_max = int(line.item_no)

        contract_product_objects = {}
        material_variant_objects = {}

        contract_materials = sap_migration_models.ContractMaterial.objects.filter(
            id__in=contract_material_ids
        )
        for contract_product in contract_materials:
            contract_product_objects[str(contract_product.id)] = contract_product
            if not material_variant_ids:
                material_variant_id = get_material_variant_by_contract_product(
                    contract_product_objects,
                    contract_product.id,
                )
                material_variant_ids.append(material_variant_id)

        for material_variant in sap_migration_models.MaterialVariantMaster.objects.filter(
                id__in=material_variant_ids
        ):
            material_variant_objects[str(material_variant.id)] = material_variant
            material_variant_codes.append(material_variant.code)
        product_group = order.product_group
        if not is_materials_product_group_matching(product_group, contract_materials, order.type):
            raise ValidationError({"product_group": ValidationError(
                "Please select the same product group to create an order",
                code=ContractCheckoutErrorCode.PRODUCT_GROUP_ERROR.value)}

            )
        if not item_no_max or product_group is None:
            product_group = contract_materials[0].mat_group_1
            update_order_product_group(order_id, product_group)
        bulk_create_lines = []
        for index, product in enumerate(products):
            contract_product_id = product.get("contract_material_id", None)
            material_variant_id = product.get("material_variant_id", None)
            if not material_variant_id:
                material_variant_id = get_material_variant_by_contract_product(
                    contract_product_objects,
                    contract_product_id
                )
            quantity = float(product.get("quantity", 0))

            i_plans = sap_migration_models.OrderLineIPlan.objects.create(
                attention_type=None,
                atp_ctp=None,
                atp_ctp_detail=None,
                block=None,
                run=None,
                iplant_confirm_quantity=None,
                item_status=None,
                original_date=None,
                inquiry_method_code=None,
                transportation_method=None,
                type_of_delivery=None,
                fix_source_assignment=None,
                split_order_item=None,
                partial_delivery=None,
                consignment=None,
                paper_machine=None,
            )
            contract_product_object = contract_product_objects.get(
                str(contract_product_id), None
            )
            material_variant_object = material_variant_objects.get(
                str(material_variant_id), None
            )

            if not contract_product_object:
                raise ValueError(
                    f"Contract product with id: {contract_product_id} not found"
                )
            if not material_variant_object:
                raise ValueError(
                    f"Material variant with id: {material_variant_id} not found"
                )
            sales_unit = contract_product_object.weight_unit or "TON" if is_default_sale_unit_from_contract(
                product_group) else "ROL"

            if is_other_product_group(product_group):
                weight = resolve_weight_common(None, material_variant_object.code, sales_unit)
            else:
                weight = resolve_weight_common(None, material_variant_object.code, "ROL")

            plant = None
            if not ProductGroup.is_default_plant_value(product_group):
                plant = contract_product_object.plant

            line = sap_migration_models.OrderLines(
                order_id=order_id,
                contract_material_id=contract_product_id,
                material_id=contract_product_object.material_id,
                material_variant_id=material_variant_id,
                quantity=quantity,
                quantity_unit=contract_product_object.quantity_unit,
                weight=weight,
                weight_unit=WeightUnitEnum.TON.value,
                total_weight=float(weight) * quantity,
                price_per_unit=contract_product_object.price_per_unit,
                total_price=net_price_calculation(product_group,
                                                  quantity,
                                                  contract_product_object.price_per_unit,
                                                  weight),
                item_no=item_no_max + (index * 10) + 10,
                original_quantity=quantity,
                net_price=net_price_calculation(
                    product_group,
                    quantity,
                    contract_product_object.price_per_unit,
                    weight
                ),
                ship_to="",
                type=OrderType.DOMESTIC.value,
                payment_term_item=contract_product_object.payment_term,
                delivery_tol_over=channel_master.over_delivery_tol
                if channel_master
                else None,
                delivery_tol_under=channel_master.under_delivery_tol
                if channel_master
                else None,
                ref_doc_it=contract_product_object.item_no,
                iplan=i_plans,
                inquiry_method=InquiryMethodType.DOMESTIC.value,
                sales_unit=sales_unit,
                price_currency=contract_product_object.currency,
                ref_doc=order.contract.code,
                plant=plant,
                prc_group_1 =contract_product_object.mat_group_1,
                shipping_mark=contract_product_object.additional_remark
            )
            if order.status_sap == ScgOrderStatusSAP.COMPLETE.value:
                line.draft = True

            bulk_create_lines.append(line)

        if len(bulk_create_lines):
            sap_migration_models.OrderLines.objects.bulk_create(bulk_create_lines)
        invalid_quantity_line_ids = validate_lines_quantity(
            order_id, contract_material_ids
        )
        if invalid_quantity_line_ids:
            lines = sap_migration_models.OrderLines.objects.filter(
                id__in=invalid_quantity_line_ids
            ).values("contract_material_id", "material_variant_id")
            raise ValueError(
                f"Total weight of products {str(combine_two_product_lists(products, list(lines)))} "
                f"are greater than remaining quantity"
            )
        sync_order_prices(order_id)
        return bulk_create_lines

    except ValidationError as e:
        transaction.set_rollback(True)
        raise e
    except Exception as e:
        transaction.set_rollback(True)
        raise ImproperlyConfigured(e)


def sync_order_prices(order_id):
    order = sap_migration_models.Order.objects.filter(id=order_id).first()
    if order:
        line_total_prices = sap_migration_models.OrderLines.all_objects.filter(
            order_id=order_id
        ).values_list("net_price", flat=True)
        total_price = sum(filter(None, line_total_prices))
        tax_percent = fn.get_tax_percent(order.sold_to.sold_to_code)
        tax_amount = float(total_price) * tax_percent
        order.total_price = float(total_price)
        order.tax_amount = tax_amount
        order.total_price_inc_tax = float(total_price) + tax_amount
        order.updated_at = datetime.now
        order.save()

    return order


def sync_order():
    """
    Sync order iPlan, SAP to order eOrdering
    @return:
    """
    sync_i_plan_data()
    sync_sap_data()


def sync_i_plan_data():
    """
    Get data from SQS of YT-65218 and sync to eOrdering Order
    @return:
    """
    queue_url, client_id, client_secret, region_name = get_sqs_config("scg.sqs_update_order")

    sqs = boto3.resource(
        'sqs',
        aws_access_key_id=client_id,
        aws_secret_access_key=client_secret,
        region_name=region_name,
    )
    queue = sqs.Queue(queue_url)
    messages = queue.receive_messages(MaxNumberOfMessages=1)
    for message in messages:
        task_logger.info(f"Start sync iPlan data: {message.body}")
        try:
            i_plan_update_order(message)
        except Exception as e:
            task_logger.error(f"Sync iPlan data error: {e}")
            message.delete()
            task_logger.info(f"Exception message removed from sync Iplan: {message.body}")
            continue
        # prevent to consume message again
        message.delete()


def order_lines_by_yt65218_message(body_message):
    order_status_request_data = body_message.get("orderStatusRequest")
    lines_data = order_status_request_data.get("orderStatusRequestLine")

    query = Q()
    for line_data in lines_data:
        order_number = line_data.get("orderNumber").zfill(10)  # SAP order number
        line_number = line_data.get("lineNumber", "").lstrip("0")  # SAP line number

        # Query order line by item_no and order so_no
        query = query | Q(item_no=line_number, order__so_no=order_number)

    qs_order_lines = sap_migration_models.OrderLines.objects.filter(query)

    dict_object_lines = {}  # Get dict line by item_no and order so_no
    for qs_line in qs_order_lines:
        so_no = str(qs_line.order.so_no)
        item_no = str(qs_line.item_no)
        dict_object_lines[f"{so_no}_{item_no}"] = qs_line

    return lines_data, dict_object_lines


def i_plan_update_order(message):
    """
    Update eOrdering order with data in SQS YT-65218
    @param message: Message of SQS
    @return:
    """
    start_time = time.time()
    logging.info("[Iplan_sync] Sync started")
    body_message = parse_log_sqs_message(message, "YT-65218")
    lines_data, dict_object_lines = order_lines_by_yt65218_message(body_message)
    dict_order_line_update_class_mark = {}
    order_lines_map = {}
    order_lines_iplan_map = {}
    orders_set = set()
    for line_data in lines_data:
        try:
            so_no = str(line_data.get("orderNumber").zfill(10))
            item_no = str(line_data.get("lineNumber"))
            object_line = dict_object_lines.get(f"{so_no}_{item_no}")
            if not object_line:
                task_logger.error(f"Order item {item_no} of order {so_no} doesn't exist!")
                continue
            order = object_line.order
            i_plan = object_line.iplan
            object_line = i_plan_convert_item_data(line_data, object_line)

            confirmed_availability_date = line_data.get("confirmedAvailabilityDate", None)
            if ProductionStatus.UNALLOCATED.value == line_data["operationStatus"]:
                logging.info(f"[Iplan_sync] so_no: {so_no},item_no: {item_no},"
                             f"DB confirmed_date: {object_line.confirmed_date},"
                             f"DB i_plant_confirm_date: {i_plan.iplant_confirm_date},"
                             f"confirmed_availability_date: {confirmed_availability_date}")
                object_line.confirmed_date = confirmed_availability_date
                i_plan.iplant_confirm_date = confirmed_availability_date

            order_type = line_data.get("orderType", None)
            if confirmed_availability_date and confirmed_availability_date is not None:
                confirm_available_to_date = datetime.strptime(confirmed_availability_date, '%Y-%m-%d').date()
                # class mark logic
                original_confirmed_date = object_line.confirmed_date
                if original_confirmed_date and line_data["operationStatus"] == ProductionStatus.COMPLETED.value:
                    class_mark_logic(object_line, confirm_available_to_date, original_confirmed_date,
                                     dict_order_line_update_class_mark)
                logging.info(f"[Iplan_sync] so_no: {so_no},item_no: {item_no},"
                             f"DB original_confirmed_date: {original_confirmed_date},"
                             f"DB i_plant_confirm_date: {i_plan.iplant_confirm_date}")
                object_line.confirmed_date = confirm_available_to_date
                i_plan.iplant_confirm_date = confirm_available_to_date

                update_attention_type_r1(object_line, confirm_available_to_date)
                update_attention_type_r4(object_line, confirm_available_to_date)

            if order_type and (AtpCtpStatus.CTP.value == order_type or AtpCtpStatus.ATP_FUTURE.value == order_type
                               or AtpCtpStatus.ATP_ON_HAND.value == order_type):
                logging.info(f"[Iplan_sync] so_no: {so_no},item_no: {item_no},"
                             f"order_type : {order_type},DB iplan order type: {i_plan.order_type}")
                i_plan.order_type = order_type
                i_plan.atp_ctp = order_type if AtpCtpStatus.CTP.value == order_type else "ATP"
                i_plan.atp_ctp_detail = order_type
            # SEO-6238
            i_plan.run = line_data.get("runCode", None)
            i_plan.block = line_data.get("blockCode", None)
            i_plan.paper_machine = line_data.get("workCentreCode", None)
            if "forAttention" in line_data:
                for_attention = line_data.get("forAttention")
                update_attention_type_r2(object_line, for_attention)
            if order not in order_lines_map:
                order_lines_map[order] = []
            order_lines_map[order].append(object_line)
            if order not in order_lines_iplan_map:
                order_lines_iplan_map[order] = []
            order_lines_iplan_map[order].append(i_plan)
            orders_set.add(order)
        except Exception as e:
            logging.info(f"[Iplan_sync] Exception while preparing items for update: {e}")

    for order in orders_set:
        try:
            with transaction.atomic():
                order_line_list = order_lines_map.get(order, None)
                order_line_iplan_list = order_lines_iplan_map.get(order, None)
                if order_line_list:
                    sap_migration_models.OrderLines.objects.bulk_update(order_line_list, fields=[
                        "assigned_quantity",
                        "production_status",
                        "item_status_en",
                        "item_status_th",
                        "attention_type",
                        "confirmed_date",
                        "class_mark",
                    ])
                if order_line_iplan_list:
                    sap_migration_models.OrderLineIPlan.objects.bulk_update(order_line_iplan_list, fields=[
                        "order_type",
                        "iplant_confirm_date",
                        "atp_ctp",
                        "atp_ctp_detail",
                        "block",
                        "run",
                        "paper_machine",
                    ])
                status_en, status_thai = update_order_status(order.id)
                logging.info(f"[Iplan_sync] order id : {order.id},"
                             f"status_en : {status_en},DB order status: {order.status}")
                order.status = status_en
                order.status_thai = status_thai
                order.save()
        except Exception as e:
            logging.info(f"[Iplan_sync] Exception while updating i_plan_sync data: {e}")
    update_class_mark_to_sap(dict_order_line_update_class_mark)
    logging.info(f"[Iplan_sync] Sync completed in :{time.time() - start_time} seconds")
    return True


def i_plan_convert_item_data(line_data, object_line):
    """
    convert IPlan data to E-ordering
    @param line_data: Data received from iPlan YT-65218
    @param object_line: Order item object in eOrdering
    @return:
    """

    assigned_quantity = None
    order_type_iplan = line_data.get("orderType") or object_line.iplan.order_type
    order_quantity = object_line.quantity
    if "reATPQuantity" in line_data:
        assigned_quantity = float(line_data.get("reATPQuantity"))
        object_line.assigned_quantity = assigned_quantity

    if operation_status := line_data.get("operationStatus"):
        # Case trigger Production operation
        object_line.production_status = operation_status

    list_status_th = IPlanOrderItemStatus.IPLAN_ORDER_LINES_STATUS_TH.value
    if order_type_iplan == AtpCtpStatus.CTP.value:
        # Case iPlan return CTP
        if operation_status == ProductionStatus.COMPLETED.value:
            if assigned_quantity == order_quantity:
                object_line.item_status_en = IPlanOrderItemStatus.FULL_COMMITTED_ORDER.value
                object_line.item_status_th = list_status_th.get(IPlanOrderItemStatus.FULL_COMMITTED_ORDER.value)
            else:
                object_line.item_status_en = IPlanOrderItemStatus.COMPLETED_PRODUCTION.value
                object_line.item_status_th = list_status_th.get(IPlanOrderItemStatus.COMPLETED_PRODUCTION.value)

        elif operation_status == ProductionStatus.UNALLOCATED.value:
            object_line.item_status_en = IPlanOrderItemStatus.PLANNING_UNALLOCATED.value
            object_line.item_status_th = list_status_th.get(IPlanOrderItemStatus.PLANNING_UNALLOCATED.value)
        elif operation_status == ProductionStatus.ALLOCATED.value:
            object_line.item_status_en = IPlanOrderItemStatus.PLANNING_ALLOCATED_NON_CONFIRM.value
            object_line.item_status_th = list_status_th.get(IPlanOrderItemStatus.PLANNING_ALLOCATED_NON_CONFIRM.value)
        elif operation_status == ProductionStatus.CONFIRMED.value:
            object_line.item_status_en = IPlanOrderItemStatus.PLANNING_CONFIRM.value
            object_line.item_status_th = list_status_th.get(IPlanOrderItemStatus.PLANNING_CONFIRM.value)
        elif operation_status == ProductionStatus.CLOSE_RUN.value:
            object_line.item_status_en = IPlanOrderItemStatus.PLANNING_CLOSE_LOOP.value
            object_line.item_status_th = list_status_th.get(IPlanOrderItemStatus.PLANNING_CLOSE_LOOP.value)
        elif operation_status == ProductionStatus.TRIMMED.value:
            object_line.item_status_en = IPlanOrderItemStatus.PLANNING_ALLOCATED_X_TRIM.value
            object_line.item_status_th = list_status_th.get(IPlanOrderItemStatus.PLANNING_ALLOCATED_X_TRIM.value)
        elif operation_status == ProductionStatus.IN_PRODUCTION.value:
            object_line.item_status_en = IPlanOrderItemStatus.PRODUCING.value
            object_line.item_status_th = list_status_th.get(IPlanOrderItemStatus.PRODUCING.value)

    elif order_type_iplan == AtpCtpStatus.ATP_FUTURE.value:
        # Case iPlan return ATP Future Stock
        if assigned_quantity == order_quantity:
            object_line.item_status_en = IPlanOrderItemStatus.FULL_COMMITTED_ORDER.value
            object_line.item_status_th = list_status_th.get(IPlanOrderItemStatus.FULL_COMMITTED_ORDER.value)
        else:
            object_line.item_status_en = IPlanOrderItemStatus.PLANNING_OUTSOURCING.value
            object_line.item_status_th = list_status_th.get(IPlanOrderItemStatus.PLANNING_OUTSOURCING.value)

    elif order_type_iplan == AtpCtpStatus.ATP_ON_HAND.value:
        # Case iPlan return ATP
        object_line.item_status_en = IPlanOrderItemStatus.FULL_COMMITTED_ORDER.value
        object_line.item_status_th = list_status_th.get(IPlanOrderItemStatus.FULL_COMMITTED_ORDER.value)

    return object_line


def sync_sap_data():
    """
    Get data from SQS of ES38 and sync to eOrdering Order
    @return:
    """
    queue_url, client_id, client_secret, region_name = get_sqs_config("scg.sqs_delivery_gi")

    sqs = boto3.resource(
        'sqs',
        aws_access_key_id=client_id,
        aws_secret_access_key=client_secret,
        region_name=region_name,
    )
    queue = sqs.Queue(queue_url)
    messages = queue.receive_messages(MaxNumberOfMessages=10)
    for message in messages:
        task_logger.info(f"Start sync SAP data: {message.body}")
        sap_update_order(message)
        # prevent to consume message again
        message.delete()


def sap_update_order(message):
    """
    Update eOrdering order with data in message of SQS ES38
    @param message: Message of SQS
    @return:
    """
    body_message = parse_log_sqs_message(message, "ES38")
    with transaction.atomic():
        try:
            lines_data = body_message.get("data")
            # for query order send to sap
            list_so_no = []
            query = Q()
            for line_data in lines_data:
                order_number = line_data.get("salesOrder")  # SAP order number
                line_number = line_data.get("salesOrderItem")  # SAP line number

                list_so_no.append(order_number)

                # Query order line by item_no and order so_no
                target_numbers = [line_number, str(int(line_number))]
                query = query | Q(item_no__in=target_numbers, order__so_no=order_number)

            qs_order_lines = sap_migration_models.OrderLines.objects.filter(query)

            dict_object_lines = {}  # Get dict line by item_no and order so_no
            for qs_line in qs_order_lines:
                so_no = str(qs_line.order.so_no)
                item_no = str(qs_line.item_no)
                dict_object_lines[f"{so_no}_{item_no}"] = qs_line

            update_lines = []
            list_orders = []
            list_order_line_update = []
            for line_data in lines_data:
                so_no = str(line_data.get("salesOrder"))
                item_no = str(line_data.get("salesOrderItem"))
                target_numbers = [item_no, str(int(item_no))]
                object_line = None
                for target_number in target_numbers:
                    object_line = dict_object_lines.get(f"{so_no}_{target_number}")
                    if object_line:
                        break
                if not object_line:
                    task_logger.error(f"[SAP ES38 sync] Order item no: {item_no} of order: {so_no} doesn't exist!")
                    continue

                list_orders.append(object_line.order)
                object_line = sap_convert_item_data(line_data, object_line)
                list_order_line_update = get_data_from_es26(object_line.order)
                logging.info(f"[Sync SAP ES38] so_no: {so_no},item_no: {item_no},"
                             f"DB delivery: {object_line.delivery} updated to :{line_data.get('delivery', '')},"
                             f"DB gi_status: {object_line.gi_status} updated to: {line_data.get('GiStatus', '')}")
                # update data from es38 message
                object_line.delivery = line_data.get("delivery")
                object_line.gi_status = line_data.get("GiStatus")
                object_line.dtr_dtp_handled = False

                # dtr dtp update
                try:
                    dtp, dtr = calculate_dtp_dtr(object_line)
                except Exception as e:
                    logging.info(f"[Sync SAP ES38] Exception occurred: {e}")
                    dtp, dtr = '', ''
                logging.info(f"[Sync SAP ES38] is DTR passed: {dtp}, is DTP passed: {dtr}")
                object_line.dtr = dtr
                object_line.dtp = dtp
                update_lines.append(object_line)
                if dtr == NOT_PASS_VALUE and dtp == NOT_PASS_VALUE:
                    add_class_mark_into_order_line(object_line, "C3", "C", 1, 4)
            if len(update_lines):
                item_status_dict = {order_line.id: (order_line.item_status_en, order_line.item_status_th) for order_line
                                    in
                                    list_order_line_update}
                for order_line in update_lines:
                    item_status_en, item_status_th = item_status_dict.get(order_line.id, (
                        order_line.item_status_en, order_line.item_status_th))
                    logging.info(f"[Sync SAP ES38] so_no: {so_no},item_no: {item_no},"
                                 f"db item_status_en: {order_line.item_status_en} updated to: {item_status_en}")
                    order_line.item_status_en = item_status_en
                    order_line.item_status_th = item_status_th
                sap_migration_models.OrderLines.objects.bulk_update(update_lines, fields=[
                    "actual_gi_date",
                    "item_status_en",
                    "item_status_th",
                    "delivery",
                    "gi_status",
                    "dtr_dtp_handled",
                    "dtr",
                    "dtp"
                ])

            need_update_orders = []
            for order in list_orders:
                if order not in need_update_orders:
                    status_en, status_thai = update_order_status(order.id)
                    logging.info(
                        f"[Sync SAP ES38] Order so_no:{order.so_no}, db status:{order.status} updated to:{status_en}")
                    order.status = status_en
                    order.status_thai = status_thai
                    need_update_orders.append(order)

            if len(need_update_orders):
                sap_migration_models.Order.objects.bulk_update(need_update_orders, fields=[
                    "status",
                    "status_thai"
                ])

            # send update to SAP ES21
            if update_lines:
                update_dtr_dtp_to_sap(list_so_no)
            return True
        except Exception as e:
            transaction.set_rollback(True)
            logging.error(e)
            return False


def sap_convert_item_data(line_data, object_line):
    """
    convert SAP data to E-ordering
    @param line_data: Data received from SAP ES38
    @param object_line: Order item object in eOrdering
    @return:
    """
    delivery = line_data.get("delivery")

    # Case update new delivery for order item
    if line_data.get("GiStatus") == DeliveryStatus.COMPLETED_DELIVERY.value:
        # actualGiDate in SAP has formatted as dd/mm/yyyy
        actual_gi_date = datetime.strptime(line_data.get("actualGiDate"), "%d/%m/%Y")
        sap_migration_models.OrderLineDeliveries.objects.create(
            order_line_id=object_line.id,
            sales_order=line_data.get("salesOrderItem"),
            sales_order_item=line_data.get("salesOrderItem"),
            delivery=delivery,
            actual_gi_date=actual_gi_date,
            gi_status=line_data.get("GiStatus"),
            sales_org=line_data.get("salesOrg"),
            distribution_channel=line_data.get("distributionChannel"),
            shipping_point=line_data.get("shippingPoint"),
        )

    # Case update cancel delivery of order item
    if line_data.get("GiStatus") == DeliveryStatus.CANCEL.value:
        sap_migration_models.OrderLineDeliveries.objects.filter(
            order_line_id=object_line.id,
            delivery=delivery,
        ).delete()

    # Update giDate to giDate of earliest remain delivery in DB or None when don't have delivery
    earliest_gi_date = sap_migration_models.OrderLineDeliveries.objects.filter(
        order_line_id=object_line.id
    ).order_by("actual_gi_date").values_list("actual_gi_date", flat=True).first()

    object_line.actual_gi_date = earliest_gi_date

    return object_line


def get_sqs_config(plugin):
    """
    Get config of SQS
    @return:
        queue_url: str
        client_id: str
        client_secret: str
        region_name: str
    """
    manager = get_plugins_manager()
    _plugin = manager.get_plugin(plugin)
    config = _plugin.config
    return config.queue_url, config.client_id, config.client_secret, config.region_name


def get_company_from_order(code):
    obj = sap_master_data_modes.SalesOrganizationMaster.objects.filter(code=code).first()
    return deepgetattr(obj, "full_name", "")


def get_po_number_from_order(order):
    if order.type == "domestic" or order.type == "customer":
        po_number = order.po_number
    else:
        po_number = order.po_no
    return po_number if po_number else ""


def prepare_order_confirmation_pdf(order):
    try:
        item_nos = []
        list_order_lines_ready = []
        list_order_lines_queue = []
        list_order_lines_cancel = []
        so_no = order.get("so_no")
        is_not_ref = order.get("is_not_ref", False)
        order_db = sap_migration_models.Order.objects.filter(so_no=so_no).first()
        if not is_not_ref:
            if not order_db:
                manager = get_plugins_manager()
                sap_fn = manager.call_api_sap_client
                es26_response = call_sap_es26(so_no=order.so_no, sap_fn=sap_fn)
                order_db = sync_export_order_from_es26(es26_response)
            if not order['contract_name']:
                order['contract_name'] = order_db.contract.project_name if order_db.contract else None
            if not order['contract_name']:
                get_contract_name_from_es14(order, so_no)
            sold_to_code = order_db.sold_to.sold_to_code if order_db.sold_to else order_db.sold_to_code
        else:
            order['contract_name'] = "This order is not ref contract"
            sold_to_code = order['order_lines'][0]['sold_to']

        for item in order["order_lines"]:
            item_nos.append(item.get("item_no"))

        for item in order["order_lines"]:
            if item.get("status", None) == SapOrderConfirmationStatus.READY_TO_SHIP.value:
                list_order_lines_ready.append(item)
            if item.get("status", None) == SapOrderConfirmationStatus.QUEUE_FOR_PRODUCTION.value:
                list_order_lines_queue.append(item)
            if item.get("status", None) == SapOrderConfirmationStatus.CANCEL.value:
                list_order_lines_cancel.append(item)
            if item.get("status", None) == SapOrderConfirmationStatus.ALL.value:
                list_order_lines_ready.append(item)
                list_order_lines_queue.append(item)
        dt_string = timezone.now().astimezone(pytz.timezone(ASIA_BANGKOK_TIMEZONE)).strftime("%d.%m.%Y/%H:%M:%S")
        date_now = timezone.now().astimezone(pytz.timezone(ASIA_BANGKOK_TIMEZONE)).strftime("%d%m%Y")
        file_name_pdf = f"OrderConfirmation_{so_no}_{date_now}"
        data_order_lines_ready = [
            {
                'item_no': item["item_no"],
                'remark': item["remark"],
                'material_description': item["material_description"],
                'qty_roll': item["order_qty"],
                'qty_of_ready_to_ship': item["confirm_qty"],
                'qty_of_queue_for_production': item["non_confirm_qty"],
                'rol': item["sale_unit"],
                'request_delivery_date': item["first_delivery_date"],
                'iplan_confirm_date': item.get("iplan_confirm_date"),
                'material_code': item["material_code"],
            }
            for item in list_order_lines_ready
        ]

        data_order_lines_queue = [
            {
                'item_no': item["item_no"],
                'remark': item["remark"],
                'material_description': item["material_description"],
                'qty_roll': item["order_qty"],
                'qty_of_ready_to_ship': item["confirm_qty"],
                'qty_of_queue_for_production': item["non_confirm_qty"],
                'rol': item["sale_unit"],
                'request_delivery_date': item["first_delivery_date"],
                'iplan_confirm_date': item.get("iplan_confirm_date"),
                'material_code': item["material_code"],
            }
            for item in list_order_lines_queue
        ]

        data_order_lines_cancel = [
            {
                'item_no': item["item_no"],
                'remark': item["remark"],
                'material_description': item["material_description"],
                'qty_roll': item["order_qty"],
                'qty_of_ready_to_ship': item["confirm_qty"],
                'qty_of_queue_for_production': item["non_confirm_qty"],
                'rol': item["sale_unit"],
                'request_delivery_date': item["first_delivery_date"],
                'iplan_confirm_date': item.get("iplan_confirm_date"),
                'material_code': item["material_code"],
            }
            for item in list_order_lines_cancel
        ]
        data_orders = {
            'sales_org_name': get_company_from_order(order["sale_org"]),
            'so_no': so_no,
            'create_order_date': order["created_date"],
            'surname': make_surname_for_send_order_email(sold_to_code),
            'sold_to_name': get_sold_to_name_and_address_from_order_for_sending_email(sold_to_code)["name"],
            'sold_to_address': get_sold_to_name_and_address_from_order_for_sending_email(sold_to_code)["address"],
            'po_no': order["po_no"],
            'payment_method_name': order["payment_term_desc"],
            'ship_to_name': get_partner_from_order_for_sending_email(order["ship_to"])["name"],
            'ship_to_address': get_partner_from_order_for_sending_email(order["ship_to"])["address"],
            'contract_no': order["contract_no"],
            'contract_name': order["contract_name"],
            "data_order_lines_ready": data_order_lines_ready,
            "data_order_lines_queue": data_order_lines_queue,
            "data_order_lines_cancel": data_order_lines_cancel,
            'created_by': order_db.created_by if order_db else "",
            'print_date_time': dt_string,
            'file_name_pdf': file_name_pdf
        }
        pdf = html_to_pdf_order_confirmation(data_orders, "order_confirm/header_confirm.html",
                                             "order_confirm/content_confirm.html", "order_confirm/footer_confirm.html",
                                             mt="-60mm", orientation="landscape")
        return f"{data_orders.get('file_name_pdf')}.pdf", pdf
    except Exception as e:
        raise ValueError(e)


def get_contract_name_from_es14(order, so_no):
    if (order["contract_no"]):
        try:
            es14_response = SapApiRequest.call_es_14_contract_detail(contract_no=order["contract_no"])
            response_data = es14_response.get("data", [])
            response_data = response_data[0]
            order['contract_name'] = response_data.get("projectName")
        except Exception as e:
            logging.exception(
                "[Order Confirmation PDF] Exception during ES14 call for contract "
                + order["contract_no"]
                + ":"
                + str(e)
            )
    else:
        logging.info(
            f"[Order Confirmation PDF] contract No came as None for Order SO No  {so_no}. Defaulting the contract Name to This order is not ref contract"
        )
        order['contract_name'] = "This order is not ref contract"


def print_pdf_order_confirmation(order):
    try:
        file_name, pdf = prepare_order_confirmation_pdf(order)
        base64_file = base64.b64encode(pdf)
        return file_name, base64_file.decode("utf-8")
    except Exception as e:
        raise ValueError(e)


def print_pdf_pending_order_report(input, info):
    try:
        response = resolve_list_of_sale_order_sap_order_pending(input, info)
        create_order_date = timezone.now().astimezone(pytz.timezone(ASIA_BANGKOK_TIMEZONE)).strftime(
            "%d.%m.%Y/%H:%M:%S")
        pdf = html_to_pdf({
            'response': response,
            'create_order_date': create_order_date
        }, "order_pending/header_pending.html", "order_pending/content_pending.html", "-10mm", "-17mm", "landscape")
        base64_file = base64.b64encode(pdf)
        date_time_format = timezone.now().astimezone(pytz.timezone(ASIA_BANGKOK_TIMEZONE))
        file_name = f'Pending_Order_Report_{date_time_format.strftime("%d%m%Y")}.pdf'
        if input.get("is_order_tracking", False):
            file_name = f'Pending_Order_Tracking_{date_time_format.strftime("%d%m%Y")}.pdf'
        return file_name, base64_file.decode("utf-8")
    except Exception as e:
        raise ValueError(e)


def download_customer_material_excel(data_input, info):
    start_time = time.time()
    try:
        logging.info(
            f"[Customer Material Mapping:Download Mappings] Request with input data: {data_input} "
            f" by user: {info.context.user}"
        )
        current_directory = os.path.dirname(os.path.abspath(__file__))
        parent_directory = get_parent_directory(current_directory, 2)
        template_file = os.path.join(parent_directory + CUSTOMER_MATERIAL_TEMPLATE_RELATIVE_PATH,
                                     CUSTOMER_MATERIAL_TEMPLATE_XLSX_FILE_NAME)

        distribution_channel_id = data_input.get("distribution_channel_id")
        sale_organization_id = data_input.get("sale_organization_id")
        sold_to_code = data_input.get("sold_to_code")
        material_response_list = resolve_list_of_customer_material(distribution_channel_id, sale_organization_id,
                                                                   sold_to_code)
        if not material_response_list:
            logger.info(f"[Customer Material Mapping:Download Mappings] No Data Found")
            return False, None, None, (f"ไม่พบข้อมูลให้ Download.")
        date_time_format = timezone.now().astimezone(pytz.timezone(ASIA_BANGKOK_TIMEZONE))
        excel_file_name = f'Customer Material Master_{sale_organization_id}_{sold_to_code}_{date_time_format.strftime(TIME_FORMAT)}'
        temporary_template_file = f'temporary_{excel_file_name}.xlsx'
        shutil.copyfile(template_file, temporary_template_file)

        wb = openpyxl.load_workbook(temporary_template_file)
        ws = wb.active
        set_column_widths(ws)

        populate_data(ws, material_response_list)
        wb.save(temporary_template_file)

        with open(temporary_template_file, "rb") as exported_file:
            base64_file = base64.b64encode(exported_file.read())
        end_time = time.time()
        logging.info(f"[Customer Material Mapping:Download Mappings] File generated successfully. Time taken: "
                    f"{end_time - start_time:.2f} seconds")
        return True, excel_file_name, base64_file.decode("utf-8"), None
    except FileNotFoundError:
        logging.error(f"[Customer Material Mapping:Download Mappings] Excel template file not found.")
        raise
    except ValueError as ve:
        raise ve
    except Exception as e:
        logging.error(f"[Customer Material Mapping:Download Mappings] An error occurred during Excel file generation.",
                     exc_info=True)
        raise ValueError("Failed to generate Excel file.") from e

def set_column_widths(ws):
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 64
    ws.column_dimensions['E'].width = 33


def populate_data(ws, material_response_list):
    for i, material in enumerate(material_response_list, start=2):
        ws[f'A{i}'] = material.sold_to_code
        ws[f'B{i}'] = material.sales_organization_code
        ws[f'C{i}'] = material.distribution_channel_code
        ws[f'D{i}'] = material.sold_to_material_code
        ws[f'E{i}'] = material.material_code
    # Set left alignment for all cells in each row
    for row_index in range(2, len(material_response_list) + 2):
        for column_index in range(1, 6):
            cell = ws.cell(row=row_index, column=column_index)
            cell.alignment = Alignment(horizontal='left')

def resolve_list_of_customer_material(distribution_channel_id, sale_organization_id, sold_to_code):
    return SoldToMaterialMaster.objects.filter(
            sales_organization_code=sale_organization_id,
            distribution_channel_code=distribution_channel_id,
            sold_to_code=sold_to_code
        ).distinct("material_code")

def download_pending_order_report_excel(data_input, info):
    try:
        field_header_column = PendingOrderFieldHeaderColumn()
        sort_columns = data_input.get('sold_to_sort')
        response = resolve_list_of_sale_order_sap_order_pending(data_input, info)
        date_time_format = timezone.now().astimezone(pytz.timezone(ASIA_BANGKOK_TIMEZONE))
        file_name = f'Pending_Order_Report_{date_time_format.strftime("%d%m%Y")}'
        if data_input.get("is_order_tracking", False):
            file_name = f'Pending_Order_Tracking_{date_time_format.strftime("%d%m%Y")}'
        row_headers = [f'อัพเดตข้อมูล: {date_time_format.strftime("%d/%m/%Y : %H:%M")}']
        temporary_file = create_file_with_headers(row_headers, ",", "xlsx")
        default_sort_field_list = ['so_no', 'item_no']

        for item in response:
            sold_to_name = get_sold_to_name_and_address_from_order_for_sending_email(item.get("sold_to"))["name"]
            row_group = [[f'Sold to: {item.get("sold_to")} {sold_to_name}']]
            for material_group in item.get("product_groups"):
                row_material_group = [material_group.get("product_group", "")]
                _row_headers = [item['th'] for item in field_header_column.find("ALL", "")]
                row_group.append(row_material_group)
                row_group.append(_row_headers)
                order_lines = material_group.get("order_lines")
                totals = material_group.get("summary")
                row_body = []
                for line in order_lines:
                    material_code = line.get('material_code', '')
                    material_des = line.get('material_description', '')
                    confirm_date = line.get("confirm_date", "")
                    order_date = line.get("order_date", "")
                    ship_to = line.get('ship_to_code', '') + " - " + line.get('ship_to_name', '')
                    _row = [
                        confirm_date,
                        order_date,
                        line.get("po_no", ""),
                        line.get("so_no", ""),
                        line.get("item_no", "").lstrip("0"),
                        material_code,
                        material_des,
                        scg_round((line.get("order_qty", 0))),
                        scg_round((line.get("pending_qty", 0))) if line.get("pending_qty", 0) or line.get("pending_qty",
                                                                                                          0) == 0 else "",
                        scg_round((line.get("atp_qty", 0))) if line.get("atp_qty", 0) or line.get("atp_qty",
                                                                                                  0) == 0 else "",
                        scg_round((line.get("ctp_qty", 0))) if line.get("ctp_qty", 0) or line.get("ctp_qty",
                                                                                                  0) == 0 else "",
                        scg_round((line.get("delivery_qty", 0))) if line.get("delivery_qty", 0) or line.get(
                            "delivery_qty", 0) == 0 else "",
                        line.get("sale_unit", ""),
                        ship_to,
                    ]
                    row_body.append(_row)
                if sort_columns:
                    for sort_column in sort_columns:
                        if item.get("sold_to") == sort_column.get("sold_to"):
                            column = field_header_column.find("en", sort_column.get("sort_field"))
                            reverse = sort_column.get("sort_type") == "DESC"
                            if column:
                                if camel_to_snake(column['en']) in default_sort_field_list:
                                    default_sort_field_list.remove(camel_to_snake(column['en']))
                                default_sort_field_list.insert(0, camel_to_snake(column['en']))
                            row_body.sort(
                                key=lambda x: multiple_sort_for_pending_order(x, default_sort_field_list, true,
                                                                              reverse=reverse),
                                reverse=reverse)
                else:
                    row_body.sort(key=lambda x: multiple_sort_for_pending_order(x, default_sort_field_list, true))
                row_group = row_group + row_body
                for total in totals:
                    sale_unit = total.get("sale_unit", "").upper() if total.get("sale_unit", "") else ""
                    summary = [
                        '',
                        '',
                        '',
                        '',
                        '',
                        '',
                        f"Total ({sale_unit})",
                        scg_round(total.get("order_qty", 0)),
                        scg_round(total.get("pending_qty", 0)),
                        scg_round(total.get("atp_qty", 0)),
                        scg_round(total.get("ctp_qty", 0)),
                        scg_round(total.get("delivery_qty", 0)),
                        sale_unit
                    ]
                    row_group.append(summary)
            row_headers.append(row_group)
            etl.io.xlsx.appendxlsx(row_group, temporary_file.name, write_header=True)

        return set_excel_column_style_for_pending_order_report(data_input, file_name, temporary_file)

    except Exception as e:
        raise ValueError(e)


def set_excel_column_style_for_pending_order_report(data_input, file_name, temporary_file):
    wb = openpyxl.load_workbook(temporary_file)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(values_only=False)):
        row_bold = False
        for cell in row:
            if isinstance(cell.value, str):
                if "Sold to" in cell.value or "Total" in cell.value:
                    cell.font = Font(bold=True)
                if "วันที่ประมาณ\nการส่งมอบ" == cell.value:
                    row_bold = True
                    set_style(ws, 'A', cell.row, width=15, alignment_style=True, height=50, auto_size=False)
                if "วันที่สั่งซื้อ\n(Create Date)" == cell.value:
                    set_style(ws, 'B', cell.row, alignment_style=True, auto_size=False)
                if "เลขที่ใบสั่งซื้อ PO\n(PO No.)" == cell.value:
                    set_style(ws, 'C', cell.row, width=18, alignment_style=True, wrap_cell_value=True,
                              min_row=cell.row + 1, min_col=3, max_col=3, auto_size=False, cell_value_alignment_horizontal_left=True)
                if "เลขที่ใบสั่งซื้อ SO\n(SO No.)" == cell.value:
                    set_style(ws, 'D', cell.row, width=18, alignment_style=True, auto_size=False)
                if "ลำดับที่\n(Item No.)"==cell.value:
                    set_style(ws, 'E', cell.row,alignment_style=True, auto_size=False)
                if "รายการสินค้า\n(Material Code)" == cell.value:
                    set_style(ws, 'F', cell.row, width=27, alignment_style=True)
                if "รายละเอียดสินค้า\n(Material Description)" == cell.value:
                    set_style(ws, 'G', cell.row, width=54, alignment_style=True)
                if "จำนวนสั่งซื้อ\n(Order Qty)" == cell.value:
                    set_style(ws, 'H', cell.row, width=14, alignment_style=False, number_format='#,##0.000', min_row=1,
                              min_col=8, max_col=8)
                if "จำนวนค้างส่ง\n(Pending Qty)" == cell.value:
                    set_style(ws, 'I', cell.row, width=14, alignment_style=False, number_format='#,##0.000', min_row=1,
                              min_col=9, max_col=9)
                if "พร้อมส่ง\n(ATP Qty)" == cell.value:
                    set_style(ws, 'J', cell.row, width=14, alignment_style=False, number_format='#,##0.000', min_row=1,
                              min_col=10, max_col=10)
                if "รอผลิต\n(CTP Qty)" == cell.value:
                    set_style(ws, 'K', cell.row, width=14, alignment_style=False, number_format='#,##0.000', min_row=1,
                              min_col=11, max_col=11)
                if "ส่งแล้ว\n(Delivery Qty)" == cell.value:
                    set_style(ws, 'L', cell.row, width=14, alignment_style=True, number_format='#,##0.000', min_row=1,
                              min_col=12, max_col=12)
                if "หน่วย\n(Unit)" == cell.value:
                    set_style(ws, 'M', cell.row, width=9, alignment_style=True)
                if "สถานที่ส่ง\n(Sold to)" == cell.value:
                    set_style(ws, 'N', cell.row, width=86, alignment_style=True)
            if row_bold:
                cell.font = Font(bold=True)
    wb.save(temporary_file.name)
    with open(temporary_file.name, "rb") as exported_file:
        if data_input.report_format == 'excel':
            excel_file = exported_file.read()
            return None, excel_file
        base64_file = base64.b64encode(exported_file.read())
    return file_name, base64_file.decode("utf-8")


def get_material_variant_by_contract_product(contract_product_objects, contract_product_id):
    """
    Get variant for material 81, 82 by contract product (contract material)
    @param contract_product_objects:
    @param contract_product_id:
    @return:
    """
    contract_product_object = contract_product_objects.get(str(contract_product_id)) or contract_product_objects.get(
        int(contract_product_id)) or None

    if not contract_product_object:
        raise ValueError(
            f"Contract product with id: {contract_product_id} not found"
        )
    variant_id = (
        sap_migration_models.MaterialVariantMaster.objects.filter(material=contract_product_object.material)
        .values_list("id", flat=True)
        .first()
    )
    if variant_id:
        return str(variant_id)
    return ""


def prepare_for_ddq_request_line(order_line, flag, order):
    parameter = change_parameter_follow_inquiry_method(order_line, order, flag)
    fmt_sold_to_code = (order.sold_to and order.sold_to.sold_to_code or order.sold_to_code or "").lstrip("0") or None
    request_line = {
        "lineNumber": order_line.item_no.lstrip("0"),
        "locationCode": fmt_sold_to_code,
        "consignmentOrder": False,
        "productCode": get_product_code(order_line),
        "requestDate": order_line.request_date.strftime("%Y-%m-%dT00:00:00.000Z") if order_line.request_date else "",
        "inquiryMethod": parameter.get("inquiry_method"),
        "quantity": str(order_line.quantity),
        "unit": "ROL",
        "transportMethod": "Truck",
        "typeOfDelivery": "E",
        "useConsignmentInventory": parameter.get("use_consignment_inventory"),
        "useProjectedInventory": parameter.get("use_projected_inventory"),
        "useProduction": parameter.get("use_production"),
        "orderSplitLogic": parameter.get("order_split_logic").upper(),
        "useInventory": parameter.get("use_inventory"),
        "singleSourcing": parameter.get("single_source"),
        "reATPRequired": parameter.get("re_atp_required"),
        "fixSourceAssignment": "",
        "consignmentLocation": order.sales_group.code,
        "requestType": flag,
        "DDQSourcingCategories": [
            {"categoryCode": order.sales_organization.code or ""},
            {"categoryCode": order.sales_group.code or ""}
        ]
    }
    return {
        "request_line": request_line,
        "parameter": parameter
    }


def prepare_params_for_i_plan(params, order_so_no, items_no, flag):
    param_order = {
        "headerCode": order_so_no.lstrip("0"),
        "autoCreate": False,
        "DDQRequestLine": []
    }
    order = sap_migration_models.Order.objects.filter(so_no=order_so_no).first()
    order_lines = (
        sap_migration_models.OrderLines.objects.filter(
            order__so_no=order_so_no,
            item_no__in=items_no,
        ).exclude(Q(item_status_en=IPlanOrderItemStatus.CANCEL.value) | Q(
            item_status_th=IPlanOrderItemStatus.IPLAN_ORDER_LINES_STATUS_TH.value.get(
                IPlanOrderItemStatus.CANCEL.value))).all()
    )
    if flag == "NEW":
        order_lines = sap_migration_models.OrderLines.objects.filter(
            order__so_no=order_so_no,
            item_no__in=items_no,
        ).first()
        result = prepare_for_ddq_request_line(order_lines, flag, order)
        request_line = result.get("request_line")
        parameter = result.get("parameter")
        param_order["DDQRequestLine"].append(
            request_line
        )
        params["DDQRequest"]["DDQRequestHeader"].append(param_order)
        return params, parameter

    for order_line in order_lines:
        param_order["DDQRequestLine"].append(
            prepare_for_ddq_request_line_delete(order_line, flag, order).get("request_line")
        )

    params["DDQRequest"]["DDQRequestHeader"].append(param_order)
    return params


def call_i_plan_request(params, *args, order=None, **kwargs):
    log_opts = {
        "orderid": order and order.id or None,
    }
    response = MulesoftApiRequest.instance(service_type=MulesoftServiceType.IPLAN.value,
                                           **log_opts).request_mulesoft_post(
        IPlanEndpoint.REQUEST_URL.value,
        params
    )
    return response


def call_es21_to_delete_cancel_orderlines(params, order=None):
    log_val = {
        "orderid": order and order.id or None,
        "order_number": order and order.so_no or None,
    }
    response = MulesoftApiRequest.instance(service_type=MulesoftServiceType.SAP.value, **log_val).request_mulesoft_post(
        SapEnpoint.ES_21.value,
        params
    )
    return response


def prepare_params_for_es21(so_no, order_lines_input):
    _order_lines = sap_migration_models.OrderLines.objects.filter(
        order__so_no=so_no,
        item_no__in=[order_line["item_no"] for order_line in order_lines_input]
    ).distinct("item_no")
    order_lines = _order_lines.in_bulk(field_name="item_no")

    params = {
        "piMessageId": str(uuid.uuid1().int),
        "salesdocumentin": so_no,
        "sapId": "newsapID",
        "testrun": False,
        "orderHeaderIn": {
            "refDoc": _order_lines.order_by("item_no").first().order.contract.code if _order_lines else ""
        },
        "orderHeaderInX": {},
        "orderItemsIn": [],
        "orderItemsInx": [],
    }

    for order_line in order_lines_input:
        order_line_db = order_lines.get(order_line.get("item_no"))
        if order_line["status"] == 'Cancel':
            params["orderItemsIn"].append(
                {
                    "itemNo": order_line_db.item_no,
                    "material": order_line_db.material_variant.code if order_line_db.material_variant else order_line_db.material.material_code,
                    "targetQty": order_line_db.quantity,
                    "salesUnit": order_line_db.sales_unit,
                    "reasonReject": "93",
                    "refDoc": order_line_db.order.contract.code,
                    "refDocIt": order_line_db.item_no,
                }
            )
            params["orderItemsInx"].append(
                {
                    "itemNo": order_line_db.item_no,
                    "updateflag": "U",
                    "reasonReject": True
                }
            )

        else:
            params["orderItemsIn"].append(
                {
                    "itemNo": order_line_db.item_no,
                    "material": order_line_db.material_variant.code if order_line_db.material_variant else order_line_db.material.material_code,
                    "targetQty": order_line_db.quantity,
                    "salesUnit": order_line_db.sales_unit,
                    "refDoc": order_line_db.contract_material.contract.code,
                    "refDocIt": order_line_db.item_no,
                }
            )
            params["orderItemsInx"].append(
                {
                    "itemNo": order_line_db.item_no,
                    "updateflag": "D",
                    "reasonReject": True
                }
            )
    return params, _order_lines


@transaction.atomic
def cancel_delete_order_lines(data, info):
    logging.info(
        f" [Domestic: Cancel/Delete] For the order {data.get('so_no', '')} FE request : {data} "
        f" by user: {info.context.user}"
    )
    success = True
    i_plan_messages_response = []
    sap_order_messages_response = []
    sap_item_messages_response = []
    sap_order_message_being_process = []
    order = sap_migration_models.Order.objects.filter(so_no=data["so_no"]).first()
    need_iplan_integration = ProductGroup.is_iplan_integration_required(order)
    items_no = [order_line["item_no"] for order_line in data["order_lines"]]
    if not need_iplan_integration:
        return cancel_delete_order_lines_no_iplan(data, info)

    logging.info(
        f" [Domestic: Cancel/Delete] For the order {order.order_no} Product group {order.product_group} "
        f" iplan call is required. Triggering the iplan cancel")

    params = {
        "DDQRequest": {
            "requestId": str(uuid.uuid1().int),
            "sender": "e-ordering",
            "DDQRequestHeader": []
        }
    }

    prepare_params_for_i_plan(params, data["so_no"], items_no, "DELETE")
    logging.info(f"[Domestic: Cancel/Delete] Calling... iplan")
    response = call_i_plan_request(params, order=order)
    logging.info(f"[Domestic: Cancel/Delete] Called iplan")
    i_plan_error_item_no, i_plan_error_messages = get_iplan_error_messages(response, True)
    logging.info(f" [Domestic: Cancel/Delete] i_plan_error_messages : {i_plan_error_messages}"
                 f"i_plan_error_item_no : {i_plan_error_item_no}")
    if i_plan_error_messages:
        success = False
        i_plan_messages_response = i_plan_error_messages
        data["order_lines"] = [item for item in data.get("order_lines", {}) if
                               item.get('item_no', None) not in i_plan_error_item_no]
        # if YT65156/plan return failed for all items
        # should not request ES21
        if len(data["order_lines"]) == 0:
            return (
                success,
                i_plan_messages_response,
                sap_order_messages_response,
                sap_item_messages_response
            )
    try:
        params_for_es21, order_lines = prepare_params_for_es21(data["so_no"], data["order_lines"])
        scgp_user = info.context.user.scgp_user
        if scgp_user and scgp_user.sap_id:
            params_for_es21["sapId"] = scgp_user.sap_id
        try:
            logging.info(f"[Domestic: Cancel/Delete] calling... ES21")
            response_es21 = call_es21_to_delete_cancel_orderlines(params_for_es21,
                                                                  order=order_lines and order_lines[0].order or None)
            logging.info(f"[Domestic: Cancel/Delete] ES21 called")
        except Exception as e:
            update_attention_type_r5(order_lines)
            logging.error(f"[Domestic: Cancel/Delete] An Exception occurred from ES-21: {str(e)}")
            success = False
            return (
                success,
                i_plan_messages_response,
                sap_order_message_being_process,
                sap_item_messages_response
            )
    except Exception as e:
        logging.error(f"[Domestic: Cancel/Delete] An Exception occurred: {str(e)}")
        success = False
        return (
            success,
            i_plan_messages_response,
            sap_order_messages_response,
            sap_item_messages_response
        )
    (
        sap_order_messages_response,
        sap_item_messages_response,
        is_being_process,
        sap_success
    ) = get_error_messages_from_sap_response_for_change_order(response_es21)

    if sap_order_messages_response and is_being_process:
        logging.info(f"[Domestic: Cancel/Delete] sap_order_error_message: {sap_order_messages_response},"
                     f"is_order_being_processed by other user: {is_being_process}")
        success = False
        update_attention_type_r5(order_lines)
        return (
            success,
            i_plan_messages_response,
            sap_order_messages_response,
            sap_item_messages_response
        )

    if sap_order_messages_response or sap_item_messages_response:
        logging.info(f"[Domestic: Cancel/Delete] sap_order_error_message:{sap_order_messages_response},"
                     f"sap_item_error_messages : {sap_item_messages_response}")
        success = False
        update_attention_type_r5(order_lines)
        return (
            success,
            i_plan_messages_response,
            sap_order_messages_response,
            sap_item_messages_response
        )
    # SAP:ES21 response is success
    update_cancel_delete_order_lines(data, info, order_lines)
    return (
        success,
        i_plan_messages_response,
        sap_order_messages_response,
        sap_item_messages_response
    )


def cancel_delete_order_lines_no_iplan(data, info):
    success = True
    i_plan_messages_response = []
    sap_order_messages_response = []
    sap_item_messages_response = []
    try:
        logging.info(f"[Domestic: Cancel/Delete] skipped iplan call and "
                     f"calling ES21 as order is related to PRODUCT_GROUP_2({ProductGroup.PRODUCT_GROUP_2.value})")
        params_for_es21, order_lines = prepare_params_for_es21(data["so_no"], data["order_lines"])
        scgp_user = info.context.user.scgp_user
        if scgp_user and scgp_user.sap_id:
            params_for_es21["sapId"] = scgp_user.sap_id
        logging.info("[Domestic: Cancel/Delete] calling... ES21")
        response_es21 = call_es21_to_delete_cancel_orderlines(params_for_es21,
                                                              order=order_lines and order_lines[0].order or None)
        logging.info("[Domestic: Cancel/Delete] called ES21")
    except Exception as e:
        logging.error(f"[Domestic: Cancel/Delete] An Exception occurred: {str(e)}")
        success = False
        return (
            success,
            i_plan_messages_response,
            sap_order_messages_response,
            sap_item_messages_response
        )

    (
        sap_order_messages_response,
        sap_item_messages_response,
        is_being_process,
        sap_success
    ) = get_error_messages_from_sap_response_for_change_order(response_es21)

    if sap_order_messages_response and is_being_process:
        logging.info(f"[Domestic: Cancel/Delete] sap_order_error_message: {sap_order_messages_response},"
                     f"is_order_being_processed by other user : {is_being_process}")
        success = False
        return (
            success,
            i_plan_messages_response,
            sap_order_messages_response,
            sap_item_messages_response
        )

    if sap_order_messages_response or sap_item_messages_response:
        logging.info(f"[Domestic: Cancel/Delete] sap_order_error_message: {sap_item_messages_response},"
                     f"sap_item_error_messages : {sap_item_messages_response}")
        success = False
        return (
            success,
            i_plan_messages_response,
            sap_order_messages_response,
            sap_item_messages_response
        )
    # SAP:ES21 response is success
    update_cancel_delete_order_lines(data, info, order_lines)
    return (
        success,
        i_plan_messages_response,
        sap_order_messages_response,
        sap_item_messages_response
    )


def update_cancel_delete_order_lines(data, info, order_lines):
    items_no = [order_line["item_no"] for order_line in data["order_lines"]]
    update_cancel_93_lines = []
    delete_lines = []
    for order_line_input in data["order_lines"]:
        for line in order_lines:
            if line.item_no == order_line_input["item_no"]:
                if order_line_input["status"] == 'Cancel':
                    update_cancel_93_lines.append(line)
                else:
                    delete_lines.append(line)
    if update_cancel_93_lines:
        _update_reject_reason_for_items(update_cancel_93_lines)
        order = sap_migration_models.Order.objects.filter(so_no=data["so_no"]).first()
        status_en, status_thai = update_order_status(order.id)
        order.status = status_en
        order.status_thai = status_thai
        order.update_by = info.context.user
        logging.info(f"[Domestic: Cancel/Delete] Item {items_no} Cancelled by  {info.context.user}")
        order.save()
    if delete_lines:
        sap_migration_models.OrderLines.objects.filter(
            id__in=[item.id for item in delete_lines]
        ).delete()
        order = sap_migration_models.Order.objects.filter(so_no=data["so_no"]).first()
        order_lines = order.orderlines_set.all()
        if not order_lines:
            order.product_group = None
        user = info.context.user
        order.update_by = user
        logging.info(f"[Domestic: Cancel/Delete] Item {items_no} Deleted by  {user}")
        order.save()



def get_data_from_es26(order):
    """
    Get data from es26 and sync to eOrdering
    @param order:
    @return:
    """
    list_order_line_update = []
    manager = get_plugins_manager()
    sap_fn = manager.call_api_sap_client
    es26_response = call_sap_es26(so_no=order.so_no, sap_fn=sap_fn)
    _order_lines_from_es26 = es26_response["data"][0].get("orderItems", None)
    if _order_lines_from_es26:
        order_text_mapping = make_order_text_mapping(es26_response["data"][0]["orderText"])
        es26_response["order_partner_mapping"] = mapping_order_partners(es26_response["data"][0]["orderPartners"])
        es26_response["order_text_mapping"] = order_text_mapping
        list_order_line_update = handle_order_lines_from_es26(order, es26_response)
    return list_order_line_update


def prepare_params_for_es21_undo(so_no, order_lines, info):
    order = sap_migration_models.Order.objects.filter(so_no=so_no).first()
    params = {
        "piMessageId": str(uuid.uuid1().int),
        "salesdocumentin": so_no,

        "testrun": False,
        "orderHeaderIn": {
            "refDoc": order.contract.code if order.contract else "",
        },
        "orderHeaderInX": {},
        "orderItemsIn": [],
        "orderItemsInx": [],

    }

    params["orderItemsIn"].append(
        {
            "itemNo": order_lines.item_no,
            "material": order_lines.material_variant.code,
            "targetQty": order_lines.quantity,
            "salesUnit": "EA" if order_lines.contract_material.material.material_group == "PK00" else order_lines.sales_unit or "ROL",
            "reasonReject": "",
            "refDoc": order_lines.ref_doc if order_lines.ref_doc else "",
            "refDocIt": order_lines.contract_material and order_lines.contract_material.item_no or "",
        }
    )
    params["orderItemsInx"].append(
        {
            "itemNo": order_lines.item_no,
            "updateflag": "U",
            "reasonReject": True
        }
    )

    return params


def call_iplan_roll_back(so_no, order_lines):
    params = {
        "DDQConfirm": {
            "requestId": str(uuid.uuid1().int),
            "sender": "e-ordering",
            "DDQConfirmHeader": [
                {
                    "headerCode": so_no,
                    "originalHeaderCode": so_no,
                    "DDQConfirmLine": []
                }
            ]
        }
    }

    ddq_confirm_line = {
        "lineNumber": order_lines.item_no,
        "originalLineNumber": order_lines.item_no,
        "status": "ROLLBACK",
        "DDQOrderInformationType": []
    }
    params["DDQConfirm"]["DDQConfirmHeader"][0]["DDQConfirmLine"].append(ddq_confirm_line)

    log_opts = {
        "orderid": order_lines.order and order_lines.order.id or None,
        "order_number": so_no,
    }
    response = MulesoftApiRequest.instance(service_type=MulesoftServiceType.IPLAN.value,
                                           **log_opts).request_mulesoft_post(
        IPlanEndpoint.IPLAN_CONFIRM_URL.value,
        params
    )
    return response


'''
SEO-4676: prepare order info for YT65156 Confirm API in the Undo of cancelled items with reason 'cancel93'
'''


def build_order_information_type(contract_no, so_no):
    order_info_types = []
    so_no = so_no.zfill(10)
    order = sap_migration_models.Order.objects.filter(Q(so_no=so_no) | Q(eo_no=so_no)).first()
    if order and order.shipping_mark:
        order_info_types.append(
            {"valueType": "ShippingMarks", "value": order.shipping_mark}
        )
    if contract_no:
        order_info_types.append(
            {"valueType": "ProformaInvoice", "value": contract_no}
        )
    ship_to_code = order.ship_to.split("-")[0].strip() if order.ship_to else None
    ship_to_partner = sap_master_data_modes.SoldToPartnerAddressMaster.objects.filter(
        partner_code=ship_to_code
    ).first()
    sold_to_code = order.sold_to.sold_to_code if order.sold_to else order.sold_to_code
    sold_to_partner = get_sold_to_partner(sold_to_code)
    if sold_to_partner:
        sold_to_name = " ".join(
            filter(
                None,
                [
                    getattr(sold_to_partner, field, "")
                    for field in ["name1", "name2", "name3", "name4"]
                ],
            )
        )
        sold_to = f"{sold_to_code} - {sold_to_name}"

        order_info_types.append({"valueType": "SoldTo", "value": sold_to})
    if ship_to_partner and ship_to_partner.country_code:
        order_info_types.append(
            {
                "valueType": "Country",
                "value": ship_to_partner.country_code,
            }
        )

    return order_info_types





'''
SEO-4676: prepare order info for YT65156 Confirm API in the Undo of cancelled items with reason 'cancel93'
'''


def build_order_information_type(contract_no, so_no):
    order_info_types = []
    so_no = so_no.zfill(10)
    order = sap_migration_models.Order.objects.filter(Q(so_no=so_no) | Q(eo_no=so_no)).first()
    if order and order.shipping_mark:
        order_info_types.append(
            {"valueType": "ShippingMarks", "value": order.shipping_mark}
        )
    order_info_types.append(
        {"valueType": "ProformaInvoice", "value": contract_no}
    )
    ship_to_code = order.ship_to.split("-")[0].strip() if order.ship_to else None
    ship_to_partner = sap_master_data_modes.SoldToPartnerAddressMaster.objects.filter(
        partner_code=ship_to_code
    ).first()
    sold_to_code = order.sold_to.sold_to_code if order.sold_to else order.sold_to_code
    sold_to_partner = get_sold_to_partner(sold_to_code)
    if sold_to_partner:
        sold_to_name = " ".join(
            filter(
                None,
                [
                    getattr(sold_to_partner, field, "")
                    for field in ["name1", "name2", "name3", "name4"]
                ],
            )
        )
        sold_to = f"{sold_to_code} - {sold_to_name}"

        order_info_types.append({"valueType": "SoldTo", "value": sold_to})
    order_info_types.append(
        {
            "valueType": "Country",
            "value": ship_to_partner.country_code if ship_to_partner else "",
        }
    )

    return order_info_types


def call_iplan_confirm_undo(so_no, order_lines, response_iplan, response_es21):
    manager = get_plugins_manager()
    confirm_quantity = 0
    order_header = {}
    order = order_lines.order
    item_no = order_lines.item_no.zfill(6)
    sap_fn = manager.call_api_sap_client
    es26_response = call_sap_es26(so_no=so_no, sap_fn=sap_fn, order_id=order.id)
    data_response = es26_response.get("data", None)
    if es26_response and data_response:
        data_response = data_response[0]
        order_header = data_response.get("orderHeaderIn", {})

    on_hand_stock = _derive_on_hand_stock_val(response_iplan)
    if not on_hand_stock:
        confirm_quantity = 0
    else:
        for order_line in response_es21["orderSchedulesOut"]:
            if order_line.get("itemNo") == item_no:
                confirm_quantity = order_line.get("confirmQuantity", 0)

    params = {
        "DDQConfirm": {
            "requestId": str(uuid.uuid1().int),
            "sender": "e-ordering",
            "DDQConfirmHeader": [
                {
                    "headerCode": so_no,
                    "originalHeaderCode": so_no,
                    "DDQConfirmLine": []
                }
            ]
        }
    }
    ddq_confirm_line = {
        "lineNumber": order_lines.item_no,
        "originalLineNumber": order_lines.item_no,
        "onHandQuantityConfirmed": str(confirm_quantity),
        "unit": "ROL",
        "status": "COMMIT",
        "DDQOrderInformationType": [
            {
                "type": "CustomInfo",
                "DDQOrderInformationItem": build_order_information_type(
                    order_header.get("contractNo", ""), so_no
                ),
            }
        ] if order_header and order_lines.order.type == "export" else [],
    }
    params["DDQConfirm"]["DDQConfirmHeader"][0]["DDQConfirmLine"].append(ddq_confirm_line)

    log_val = {
        "orderid": order and order.id or None,
        "order_number": order and order.so_no or None,
    }
    response = MulesoftApiRequest.instance(service_type=MulesoftServiceType.IPLAN.value,
                                           **log_val).request_mulesoft_post(
        IPlanEndpoint.IPLAN_CONFIRM_URL.value,
        params
    )
    return response


def call_es21_to_undo_orderlines(params, order: sap_migration_models.Order = None):
    log_val = {
        "orderid": order and order.id or None,
        "order_number": order and order.so_no or None,
    }
    response = MulesoftApiRequest.instance(service_type=MulesoftServiceType.SAP.value, **log_val).request_mulesoft_post(
        SapEnpoint.ES_21.value,
        params
    )
    return response


@transaction.atomic
def undo_order_lines(data, info):
    success = True
    i_plan_messages_response = []
    sap_order_messages_response = []
    sap_item_messages_response = []
    manager = get_plugins_manager()
    params = {
        "DDQRequest": {
            "requestId": str(uuid.uuid1().int),
            "sender": "e-ordering",
            "DDQRequestHeader": []
        }
    }
    # for one order line
    so_no = data["so_no"]
    item_no = data["item_no"]
    logging.info(
        f"[Domestic: Undo order lines] for order so_no : {so_no}  FE request: {data} by user {info.context.user}")
    params, parameter = prepare_params_for_i_plan(params, so_no, item_no, "NEW")
    order = sap_migration_models.Order.objects.filter(so_no=so_no).first()
    order_line = sap_migration_models.OrderLines.objects.select_related("iplan", "material_variant").filter(
        order__so_no=so_no,
        item_no__in=item_no
    ).first()
    need_iplan_integration = ProductGroup.is_iplan_integration_required(
        order)
    params_for_es21 = prepare_params_for_es21_undo(so_no, order_line, info)
    scgp_user = info.context.user.scgp_user
    if scgp_user and scgp_user.sap_id:
        params_for_es21["sapId"] = scgp_user.sap_id
    response = {}
    logging.info(f"[Domestic: Undo order lines] is order {so_no} related to PRODUCT_GROUP_1 : {need_iplan_integration}")
    if need_iplan_integration:
        logging.info("[Domestic: Undo order lines] calling....Iplan")
        response = call_i_plan_request(params, manager, order=order)
        logging.info("[Domestic: Undo order lines] called Iplan")
        i_plan_error_messages = get_iplan_error_messages(response)
        logging.info(f"[Domestic: Undo order lines] i_plan_error_messages: {i_plan_error_messages}")
        if i_plan_error_messages:
            success = False
            i_plan_messages_response = i_plan_error_messages
            return (
                success,
                i_plan_messages_response,
                sap_order_messages_response,
                sap_item_messages_response
            )
    logging.info("[Domestic: Undo order lines] calling......ES21")
    response_es21 = call_es21_to_undo_orderlines(params_for_es21, order)
    logging.info("[Domestic: Undo order lines] called ES21")

    (
        sap_order_messages_response,
        sap_item_messages_response,
        is_being_process,
        sap_success
    ) = get_error_messages_from_sap_response_for_change_order(response_es21)

    if sap_order_messages_response or sap_item_messages_response:
        logging.info(f"[Domestic: Undo order lines] sap_order_error_message: {sap_order_messages_response},"
                     f"sap_item_error_messages: {sap_item_messages_response}")
        success = False
        if need_iplan_integration:
            logging.info("[Domestic: Undo order lines] called iplan_roll_back as ES21 is failed")
            call_iplan_roll_back(data["so_no"].lstrip("0"), order_line)
        return (
            success,
            i_plan_messages_response,
            sap_order_messages_response,
            sap_item_messages_response
        )
    if need_iplan_integration:
        logging.info("[Domestic: Undo order lines] calling... iplan_confirm_commit")
        response_i_plan_confirm = call_iplan_confirm_undo(data["so_no"].lstrip("0"), order_line, response,
                                                          response_es21)
        logging.info("[Domestic: Undo order lines] called iplan_confirm")
        i_plan_error_messages = get_iplan_error_messages(response_i_plan_confirm)
        logging.info(f"[Domestic: Undo order lines] i_plan_confirm_undo_error_messages: {i_plan_error_messages}")
        if i_plan_error_messages:
            update_attention_type_r5([order_line])
            i_plan_messages_response = i_plan_error_messages
            success = False
            return (
                success,
                i_plan_messages_response,
                sap_order_messages_response,
                sap_item_messages_response
            )
        update_iplan_for_order_line(so_no, response, parameter)
    else:
        update_order_lines_skipped_iplan(order, item_no)
    status_en, status_thai = update_order_status(order.id)
    logging.info(f"[Domestic: Undo order lines] DB order {order.so_no} status : {order.status} updated to {status_en}")
    order.status = status_en
    order.status_thai = status_thai
    order.save()
    logging.info(
        f"[Domestic: Undo order lines] Undo item {item_no} of order with SoNo : {order.so_no} is completed successfully"
    )
    return (
        success,
        i_plan_messages_response,
        sap_order_messages_response,
        sap_item_messages_response
    )


def update_order_lines_skipped_iplan(order, item_no):
    order_lines = sap_migration_models.OrderLines.objects.filter(order=order)
    dict_order_lines = {instance.item_no: instance for instance in order_lines}
    order_line = dict_order_lines.get(item_no[0])
    e_ordering_order_lines = []
    order_line.item_status_en = EorderingItemStatusEN.ITEM_CREATED.value
    order_line.item_status_th = EorderingItemStatusTH.ITEM_CREATED.value
    e_ordering_order_lines.append(order_line)
    sap_migration_models.OrderLines.objects.bulk_update(
        e_ordering_order_lines,
        fields=[
            "item_status_en",
            "item_status_th",
        ])


@transaction.atomic
def update_iplan_for_order_line(so_no, response, parameter):
    try:
        e_ordering_order_lines = []
        new_order_lines = []
        i_plan_order_lines = response.get("DDQResponse").get("DDQResponseHeader")[0].get("DDQResponseLine")
        order = sap_migration_models.Order.objects.filter(so_no=so_no).first()
        order_lines = sap_migration_models.OrderLines.objects.filter(order=order)
        dict_order_lines = {instance.item_no: instance for instance in order_lines}
        if not dict_order_lines:
            return []

        max_item_no = max([int(line.item_no) for line in order_lines])

        e_ordering_order_lines_i_plan = []
        for i_plan_line in i_plan_order_lines:
            is_split_item = False
            line_id = i_plan_line.get("lineNumber").lstrip("0")
            e_ordering_order_line = dict_order_lines.get(line_id)
            if not e_ordering_order_line:
                is_split_item = True
                max_item_no += 10
                e_ordering_order_line = deepcopy(dict_order_lines.get(str(floor(float(line_id)))))
                e_ordering_order_line.pk = None
                e_ordering_order_line.iplan = None
                e_ordering_order_line.item_no = str(max_item_no)

            i_plan_on_hand_stock = i_plan_line.get("onHandStock")
            assigned_quantity = 0

            i_plan_operation = {}
            if i_plan_operations := (i_plan_line.get("DDQResponseOperation") or None):
                i_plan_operation = i_plan_operations[0]

            i_plan_confirm_quantity = i_plan_line.get("quantity", 0)
            i_plan_confirmed_date = i_plan_line.get("dispatchDate") or None
            i_plan_plant = i_plan_line.get("warehouseCode", None)
            block = i_plan_operation.get("blockCode", None)
            run = i_plan_operation.get("runCode", None)
            item_status = i_plan_line.get("status", None)
            original_date = i_plan_line.get("dispatchDate") or None
            paper_machine = i_plan_operation.get("workCentreCode", None)
            order_type = i_plan_line.get("orderType", None)
            plant = i_plan_line.get("warehouseCode", None)
            item_no = i_plan_line.get("lineNumber", None)
            on_hand_stock = i_plan_line.get("onHandStock", None)
            # Update order line i-plan table
            e_ordering_order_line_i_plan = e_ordering_order_line.iplan

            atp_ctp = handle_atp_ctp(i_plan_operations, i_plan_on_hand_stock)
            if i_plan_on_hand_stock:
                assigned_quantity = i_plan_line.get("quantity", 0)

            if not e_ordering_order_line_i_plan:
                e_ordering_order_line_i_plan = (
                    sap_migration_models.OrderLineIPlan.objects.create(
                        atp_ctp=atp_ctp,
                        iplant_confirm_quantity=i_plan_confirm_quantity,
                    )
                )
                e_ordering_order_line.iplan = e_ordering_order_line_i_plan

            e_ordering_order_line_i_plan.atp_ctp = atp_ctp
            e_ordering_order_line_i_plan.iplant_confirm_quantity = i_plan_confirm_quantity
            e_ordering_order_line_i_plan.atp_ctp_detail = order_type
            e_ordering_order_line_i_plan.block = block
            e_ordering_order_line_i_plan.run = run
            e_ordering_order_line_i_plan.item_status = item_status
            e_ordering_order_line_i_plan.original_date = original_date
            e_ordering_order_line_i_plan.partial_delivery = "false"
            e_ordering_order_line_i_plan.paper_machine = paper_machine
            e_ordering_order_line_i_plan.inquiry_method_code = parameter.get("inquiry_method")
            e_ordering_order_line_i_plan.transportation_method = "Truck"
            e_ordering_order_line_i_plan.type_of_delivery = IPlanTypeOfDelivery.EX_MILL.value
            e_ordering_order_line_i_plan.fix_source_assignment = e_ordering_order_line.plant or ""
            e_ordering_order_line_i_plan.split_order_item = parameter.get("order_split_logic")
            e_ordering_order_line_i_plan.iplant_confirm_date = i_plan_confirmed_date
            e_ordering_order_line_i_plan.consignment = "false"
            e_ordering_order_line_i_plan.use_inventory = parameter.get("use_inventory")
            e_ordering_order_line_i_plan.use_consignment_inventory = parameter.get("use_consignment_inventory")
            e_ordering_order_line_i_plan.use_projected_inventory = parameter.get("use_projected_inventory")
            e_ordering_order_line_i_plan.use_production = parameter.get("use_production")
            e_ordering_order_line_i_plan.single_source = parameter.get("single_source")
            e_ordering_order_line_i_plan.re_atp_required = parameter.get("re_atp_required")
            e_ordering_order_line_i_plan.request_type = "NEW"
            e_ordering_order_line_i_plan.order_type = order_type
            e_ordering_order_line_i_plan.plant = plant
            e_ordering_order_line_i_plan.item_no = item_no
            e_ordering_order_line_i_plan.on_hand_stock = on_hand_stock

            e_ordering_order_lines_i_plan.append(e_ordering_order_line_i_plan)

            # Update order line table
            e_ordering_order_line.i_plan_on_hand_stock = i_plan_on_hand_stock
            e_ordering_order_line.i_plan_operations = i_plan_operations
            e_ordering_order_line.confirmed_date = i_plan_confirmed_date
            e_ordering_order_line.quantity = i_plan_line.get("quantity")
            e_ordering_order_line.plant = i_plan_plant
            e_ordering_order_line.original_item_no = line_id
            e_ordering_order_line.return_status = item_status
            item_status_value = IPlanOrderItemStatus.FULL_COMMITTED_ORDER.value if (
                    AtpCtpStatus.ATP_ON_HAND.value == order_type) else IPlanOrderItemStatus.ITEM_CREATED.value
            if order_type == AtpCtpStatus.ATP_FUTURE.value:
                item_status_value = IPlanOrderItemStatus.PLANNING_OUTSOURCING.value
            e_ordering_order_line.item_status_en = item_status_value
            e_ordering_order_line.item_status_th = IPlanOrderItemStatus.IPLAN_ORDER_LINES_STATUS_TH.value.get(
                item_status_value
            )
            e_ordering_order_line.assigned_quantity = assigned_quantity
            if is_split_item:
                new_order_lines.append(e_ordering_order_line)
            else:
                e_ordering_order_lines.append(e_ordering_order_line)

        sap_migration_models.OrderLines.objects.bulk_update(
            e_ordering_order_lines,
            fields=[
                "i_plan_on_hand_stock",
                "i_plan_operations",
                "iplan",
                "plant",
                "confirmed_date",
                "quantity",
                "original_item_no",
                "return_status",
                "assigned_quantity",
                "item_status_en",
                "item_status_th",
            ],
        )
        if len(new_order_lines) > 0:
            sap_migration_models.OrderLines.objects.bulk_create(new_order_lines)

        if len(e_ordering_order_lines_i_plan):
            sap_migration_models.OrderLineIPlan.objects.bulk_update(
                e_ordering_order_lines_i_plan,
                fields=[
                    "atp_ctp",
                    "iplant_confirm_quantity",
                    "atp_ctp_detail",
                    "block",
                    "run",
                    "item_status",
                    "original_date",
                    "partial_delivery",
                    "paper_machine",
                    "inquiry_method_code",
                    "transportation_method",
                    "type_of_delivery",
                    "fix_source_assignment",
                    "split_order_item",
                    "consignment",
                    "use_inventory",
                    "use_consignment_inventory",
                    "use_projected_inventory",
                    "single_source",
                    "use_production",
                    "re_atp_required",
                    "request_type",
                    "order_type",
                    "iplant_confirm_date",
                    "plant",
                    "on_hand_stock",
                    "item_no",
                ],
            )

        mock_confirmed_date(order)
    except Exception as e:
        transaction.set_rollback(True)
        if isinstance(e, ConnectionError):
            raise ValueError("Error Code : 500 - Internal Server Error")
        raise ImproperlyConfigured(e)


def update_attention_type_r1(line, confirm_output_date):
    attention_type = line.attention_type or "R1"
    '''
    SEO-4434: 
        Flag R1: request date < confirmed date or request date < confirm availability date
        Unflag R1: request >= confirmed date or request >= confirm availability date
    '''
    if line.request_date < line.confirmed_date or line.request_date < confirm_output_date:
        attention_type = ','.join(sorted(attention_type.replace("R1", "").split() + ['R1'])).strip(", ")
    else:
        attention_type = attention_type.replace("R1", "").strip(", ")
    line.attention_type = attention_type


def update_attention_type_r2(line, for_attention):
    attention_type = line.attention_type or ""
    attention_list = [att.strip() for att in attention_type.split(",") if att.strip()]
    if for_attention:
        if "R2" not in attention_list:
            attention_list.append("R2")
    else:
        if "R2" in attention_list:
            attention_list.remove("R2")
    line.attention_type = ", ".join(sorted(attention_list)).strip(", ")


def _parse_datetime(date_str):
    if not date_str:
        return None
    format_dates = ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%d%m%Y"]
    for fmt in format_dates:
        try:
            target_date = datetime.strptime(date_str, fmt).date()
            return target_date
        except Exception as e:
            logging.warning(e)
            continue
    return None


def update_attention_type_r4(line, confirm_output_date):
    attention_type = line.attention_type or ""
    attention_list = [att.strip() for att in attention_type.split(",") if att.strip()]
    etd = _parse_datetime(line.order.etd)
    if etd:
        if (line.confirmed_date and line.confirmed_date > etd) or (confirm_output_date and confirm_output_date > etd):
            if "R4" not in attention_list:
                attention_list.append("R4")
        else:
            if "R4" in attention_list:
                attention_list.remove("R4")
    line.attention_type = ", ".join(sorted(attention_list)).strip(", ")


def prepare_for_ddq_request_line_delete(order_line, flag, order):
    parameter = change_parameter_follow_inquiry_method(order_line, order)
    fmt_sold_to_code = (order.sold_to and order.sold_to.sold_to_code or order.sold_to_code or "").lstrip("0") or None
    request_line = {
        "lineNumber": order_line.item_no.lstrip("0"),
        "locationCode": fmt_sold_to_code,
        "productCode": get_product_code(order_line),
        "requestDate": order_line.request_date.strftime("%Y-%m-%dT00:00:00.000Z") if order_line.request_date else "",
        "inquiryMethod": parameter.get("inquiry_method"),
        "quantity": str(order_line.quantity),
        "transportMethod": "Truck",
        "typeOfDelivery": "E",
        "singleSourcing": parameter.get("single_source"),
        "requestType": flag,
        "unit": "ROL"
    }
    return {
        "request_line": request_line,
        "parameter": parameter
    }


def update_item_status_for_special_container_items(order_line):
    order_line.item_status_en = IPlanOrderItemStatus.ITEM_CREATED.value
    order_line.item_status_th = IPlanOrderItemStatus.IPLAN_ORDER_LINES_STATUS_TH.value.get(
        IPlanOrderItemStatus.ITEM_CREATED.value
    )
    order_line.save()


@transaction.atomic
def parse_log_sqs_message(message, sqs_type):
    body_message = "{}"  # default, update later
    try:
        body = json.loads(message.body)
        body_message = parse_json_if_possible(body.get("Message"))
    except Exception as e:
        logging.error("Error get log sqs message: %s", str(e))
    sap_migration_models.SqsLog.objects.create(
        name=sqs_type,
        message=body_message,
        created_at=datetime.now(pytz.timezone(ASIA_BANGKOK_TIMEZONE))
    )
    return body_message


def _derive_on_hand_stock_val(response_iplan):
    if response_iplan["DDQResponse"]["DDQResponseHeader"] and \
            response_iplan["DDQResponse"]["DDQResponseHeader"][0]["DDQResponseLine"]:
        return response_iplan["DDQResponse"]["DDQResponseHeader"][0]["DDQResponseLine"][0]["onHandStock"]
    return True


def _convert_to_datetime(date_str):
    return datetime.strptime(date_str, "%d/%m/%Y")
