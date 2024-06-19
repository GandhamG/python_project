import base64
import logging
import os

import openpyxl
import petl as etl
import pytz
from django.core.exceptions import ImproperlyConfigured
from django.template.loader import get_template
from django.utils import timezone
from openpyxl.styles import Font, PatternFill

from saleor.csv.utils.export import create_file_with_headers
from scg_checkout.graphql.helper import get_parent_directory
from scg_checkout.graphql.implementations.customer_materials import EXCEL_CONTENT_TYPE
from scgp_require_attention_items.graphql.helper import set_style

EXCEL_UPLOAD_TEMPLATE_XLSX_FILE_NAME = "Template Excel Upload.xlsx"
EXCEL_UPLOAD_TEMPLATE_RELATIVE_PATH = "/templates/excel_upload/"
EXCEL_ATTACHMENT_ROW_HEADER_NAME = [
    "SO Number",
    "Material Code/ Customer Material",
    "Material Description",
    "Error Message",
    "Line",
]


def export_excel_upload_template_file(info):
    try:
        logging.info(
            f"[Excel Upload : Download Template] requested by user : '{info.context.user}'"
        )
        current_directory = os.path.dirname(os.path.abspath(__file__))
        parent_directory = get_parent_directory(current_directory, 1)
        absolute_path = os.path.join(
            parent_directory + EXCEL_UPLOAD_TEMPLATE_RELATIVE_PATH,
            EXCEL_UPLOAD_TEMPLATE_XLSX_FILE_NAME,
        )
        logging.info(f"[Excel Upload: Download Template] file path: '{absolute_path}'")
        with open(absolute_path, "rb") as template_file:
            file_contents = template_file.read()
        if file_contents:
            return (
                EXCEL_UPLOAD_TEMPLATE_XLSX_FILE_NAME,
                EXCEL_CONTENT_TYPE,
                base64.b64encode(file_contents).decode("utf-8"),
            )
        else:
            raise FileNotFoundError(
                f"{EXCEL_UPLOAD_TEMPLATE_XLSX_FILE_NAME} File doesn't exist"
            )
    except FileNotFoundError as error:
        logging.error(f"[Excel Upload: Download Template] Failed with: {error}")
        raise error
    except Exception as e:
        logging.error(f"[Excel Upload] Download Template Failed with: {e}")
        raise ImproperlyConfigured("Internal Server Error.")


def send_email_for_excel_upload_orders(manager, param=None):
    try:
        logging.info("[Excel Upload]- Going to send email for excel upload")
        # prepare email subject
        double_quotes = '"'
        file_name = f"{double_quotes}{param['file']}{double_quotes}"

        file_name = f"{double_quotes}{param['file']}{double_quotes}"
        param["upload_date_time"] = (
            param["upload_date_time"]
            .astimezone(pytz.timezone("Asia/Bangkok"))
            .strftime("%d-%m-%Y %H:%M:%S")
        )
        logging.info(f" time printed : {param['upload_date_time']}")
        if param["total_fail_order"] == 0 and param["total_partial_success_order"] == 0:
            subject = f"Success {param['sale_org_short_name']} Excel Upload {file_name} {param['upload_date_time']}"
        else:
            subject = f"Fail {param['sale_org_short_name']} Excel Upload {file_name} {param['upload_date_time']}"
        # prepare email body data
        template_data = {
            "sale_org_info": f"{param['sale_org_code']} {param['sale_org_name']}",
            "file_name": param["file"],
            "save_sap_date_time": param["save_sap_date_time"],
            "uploader_name": param["uploader"],
            "total_number_of_order": f'{param["total_number_of_order"]} Order',
            "fail_orders": f'{param["total_fail_order"]} Order',
            "partial_success_orders": f'{param["total_partial_success_order"]} Order',
            "success_orders": f'{param["total_success_orders"]} Order',
        }
        template = get_template("excel_upload_email_body.html")
        content = template.render(template_data)
        if param["total_fail_order"] == 0 and param["total_partial_success_order"] == 0:
            # send without attachment
            manager.scgp_send_order_confirmation_email(
                "scg.email",
                recipient_list=param["to"],
                subject=subject,
                template="order_confirmation_email.html",
                template_data={"content": content},
                cc_list=param["cc"],
                pdf_file=[],
                file_name_pdf="",
            )
        else:
            # prepare attachment file of type excel if there is fail order or partial order
            attachment_file_name = f"Fail_ExcelUpload_{timezone.now().astimezone(pytz.timezone('Asia/Bangkok')).strftime('%Y%m%d%H%M%S')}"
            temporary_file = create_file_with_headers([], ",", "xlsx")
            for item in param["order_status_details"]:
                row_group = [
                    [
                        f'Sold-to:{item.get("sold_to_code")} - {item.get("sold_to_name")} PO No:{item.get("po_number")}'
                    ],
                    EXCEL_ATTACHMENT_ROW_HEADER_NAME,
                ]
                row_body = []
                for line in item.get("lines"):
                    _row = [
                        line.get("so_no"),
                        line.get("mat_code"),
                        line.get("mat_desc"),
                        line.get("error"),
                        line.get("line_no"),
                    ]
                    row_body.append(_row)
                row_group = row_group + row_body

                etl.io.xlsx.appendxlsx(
                    row_group, temporary_file.name, write_header=True
                )
            apply_style_to_excel_file(temporary_file)
            with open(temporary_file.name, "rb") as exported_file:
                excel_file = exported_file.read()

            manager.scgp_send_email_with_excel_attachment(
                "scg.email",
                recipient_list=param["to"],
                subject=subject,
                template="order_confirmation_email.html",
                template_data={"content": content},
                cc_list=param["cc"],
                excel_file=excel_file,
                file_name_excel=attachment_file_name,
            )
        logging.info("[Excel Upload]- Email sent successfully")
    except Exception as e:
        raise ValueError(e)


def apply_style_to_excel_file(temporary_file):
    wb = openpyxl.load_workbook(temporary_file)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(values_only=False)):
        logging.info(f"{i}")
        row_bold = False
        fill_color = False
        for cell in row:
            if isinstance(cell.value, str):
                if "Sold-to:" in cell.value:
                    ws.column_dimensions[
                        ws.cell(row=cell.row, column=1).column_letter
                    ].width = 40
                if "Sold-to:" in cell.value or "SO Number" in cell.value:
                    row_bold = True
                    fill_color = True

                if "SO Number" in cell.value:
                    set_style(
                        ws,
                        "A",
                        cell.row,
                        width=18,
                        alignment_style=True,
                        auto_size=False,
                    )

                if "Material Code/ Customer Material" in cell.value:
                    set_style(
                        ws,
                        "B",
                        cell.row,
                        width=35,
                        alignment_style=True,
                        auto_size=False,
                    )
                if "Material Description" in cell.value:
                    set_style(
                        ws,
                        "c",
                        cell.row,
                        width=40,
                        alignment_style=True,
                        auto_size=False,
                    )
                if "Error Message" in cell.value:
                    set_style(
                        ws,
                        "D",
                        cell.row,
                        width=40,
                        alignment_style=True,
                        auto_size=False,
                    )
                if "Line" in cell.value:
                    set_style(
                        ws,
                        "E",
                        cell.row,
                        width=16,
                        min_row=cell.row + 1,
                        min_col=cell.column,
                        max_col=cell.column,
                        alignment_style=True,
                        auto_size=False,
                        cell_value_alignment_horizontal_left=True,
                    )

            if fill_color:
                cell.fill = PatternFill(
                    start_color="d4ebfb", end_color="d4ebfb", fill_type="solid"
                )

            if row_bold:
                cell.font = Font(bold=True)
    wb.save(temporary_file.name)
