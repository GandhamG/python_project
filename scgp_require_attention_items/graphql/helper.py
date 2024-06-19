import base64
import json
import logging
import uuid
import re
from datetime import datetime

import openpyxl
import petl as etl
from django.db import transaction
from django.db.models import QuerySet
from django.utils import timezone as django_tz
from django.conf import settings
from openpyxl.styles import Alignment
from pytz import timezone

from common.helpers import snake_to_camel, DateHelper
from common.product_group import ProductGroup, ProductGroupDescription
from saleor.account.models import User
from common.enum import MulesoftServiceType
from common.mulesoft_api import MulesoftApiRequest
import sap_master_data.models
from saleor.csv.utils.export import create_file_with_headers
from sap_master_data import models as master_data_models
from scg_checkout.graphql.enums import PendingOrderFieldHeaderColumn, ALIGNMENT_STYLE, ALIGNMENT_VERTICAL_TOP
from sap_migration import models as sap_migrations_models
from sap_migration.graphql.enums import InquiryMethodType
from scgp_cip.common.constants import BOM_FLAG_TRUE_VALUE
from scgp_cip.dao.order_line.order_line_repo import OrderLineRepo
from scgp_cip.service.helper.order_line_helper import is_bom_parent, is_bom_child, sorted_and_merged_order_line_list, \
    separate_parent_and_bom_order_lines
from scgp_export.graphql.enums import SapEnpoint, ItemCat, IPlanEndPoint
from scgp_require_attention_items import models as scgp_require_attention_items_model
from utils.enums import IPlanInquiryMethodCode

KG_TO_TON = 1000
INTERNAL_SERVER_ERROR = "Internal server error"
PLANT_MAPPING = {
    "7521": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "โรงงานบ้านโป่ง"
    },
    "7523": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "โรงงานวังศาลา"
    },
    "7524": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "กระดาษรุ่งโรจน์ (BAG)"
    },
    "7525": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "กระดาษรุ่งโรจน์ (SLIT)"
    },
    "7526": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "รวมไทยบรรจุภัณฑ์"
    },
    "7527": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "SKK สระบุรี  (BAG)"
    },
    "7531": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "บ้านโป่ง"
    },
    "7533": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "บ้านโป่ง-วังศาลา"
    },
    "7534": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "Wangsala/ASI"
    },
    "7536": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "บ้านโป่ง-สุขสวัสดิ์"
    },
    "7537": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "บ้านโป่ง-ปราจีนบุรี"
    },
    "753A": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "บ้านโป่ง-เคเลเซอร์"
    },
    "7546": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "BP-ESC LaemChabang์"
    },
    "754A": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "BP-LaemChabang-16์"
    },
    "754C": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "SGI สระบุรี-17"
    },
    "754D": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "SCG สงขลา-17"
    },
    "754E": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "TGP ชลบุรี-17"
    },
    "7550": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "โรงงานบ้านโป่ง-16"
    },
    "7561": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "โรงงานวังศาลา"
    },
    "7563": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "โรงงานบ้านโป่ง"
    },
    "7564": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "โรงงานท่าม่วง"
    },
    "7565": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "โรงงานปราจีนบุรี"
    },
    "7566": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "โรงงานวังศาลา-(Purc)"
    },
    "7567": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "โรงงานวังศาลา-บ้านโป่ง (Purc)"
    },
    "7568": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "โรงงานวังศาลา-16"
    },
    "7581": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "โรงงานวังศาลา-Paper Board"
    },
    "7583": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "โรงงานวังศาลา-Paper Board-TCP"
    },
    "7584": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "วังศาลา-พรีซิสชั่น"
    },
    "7585": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "วังศาลา-SGI สระบุรี"
    },
    "7586": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "วังศาลา-SGI สงขลาี"
    },
    "7587": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "Paper Board-สำโรงี"
    },
    "7588": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "วังศาลา-TGP ชลบุรี"
    },
    "PO41": {
        "name1": "บริษัท ไทยเคนเปเปอร์ จำกัด (มหาชน)",
        "name2": "คลังปราจีนบุรี"
    },
    "PO31": {
        "name1": "บริษัท ไทยเคนเปเปอร์ จำกัด (มหาชน)",
        "name2": "คลังกาญจนบุรี"
    },
    "PO32": {
        "name1": "บริษัท ไทยเคนเปเปอร์ จำกัด (มหาชน)",
        "name2": "คลังปราจีนบุรี"
    },
    "PO42": {
        "name1": "บริษัท ไทยเคนเปเปอร์ จำกัด (มหาชน)",
        "name2": "คลังกาญจนบุรี-ท่าม่วง"
    },
    "PO51": {
        "name1": "บริษัท ไทยเคนเปเปอร์ จำกัด (มหาชน)",
        "name2": "คลังกาญจนบุรี"
    },
    "754B": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "ESC-17"
    },
    "7551": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "โรงงานบ้านโป่ง-17"
    },
    "7569": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "โรงงานวังศาลา-17"
    },
    "7535": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "โรงงานบ้านโป่ง"
    },
    "7552": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "IBB บ้านโป่ง"
    },
    "756A": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "IBB วังศาลา"
    },
    "752A": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "[SD] SKIC ดิศรณ์ บรรจุภัณฑ์"
    },
    "752B": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "[SD] SKIC เอส เอ็น พี เปเปอร์"
    },
    "752C": {
        "name1": "บริษัท สยามคราฟท์อุตสาหกรรม จำกัด",
        "name2": "[SD] SKIC ยูนิค เปเปอร์์"
    }
}
DATE_FORMAT_ISO = "%d/%m/%Y"
DATE_FORMAT = "%Y-%m-%d"
SAP_STATUS_MAPPING = {
    "A": "Open",
    "B": "Being processed",
    "C": "Completed",
    "": "May not occur"
}
CURRENCY_FORMAT_CONSTANT = '.3f'


@transaction.atomic
def append_field(data_field, value):
    if not data_field:
        return value
    if value in data_field:
        return data_field
    data_field += ", " + value
    return data_field


@transaction.atomic
def sort_values(data_field, min_range, max_range, prefix):
    result = ""
    for num in range(min_range, max_range + 1):
        if str(num) in data_field:
            result += prefix + str(num) + ", "
    return result[:-2]


@transaction.atomic
def add_class_mark_into_order_line(object_line, value, prefix, min_range, max_range):
    if not value:
        return
    value_class_mark = object_line.class_mark
    value_class_mark = append_field(value_class_mark, value)
    # object_line.remark = sort_values(value_class_mark, min_range, max_range, prefix)
    object_line.class_mark = value_class_mark
    object_line.save()


def stamp_class_mark_for_original_confirmed_date_from_i_plan(original_confirmed_date, order_line):
    try:
        if order_line.confirmed_date > original_confirmed_date:
            require_attention = scgp_require_attention_items_model.RequireAttention.objects \
                .filter(id=order_line.id, type=order_line.type).first()
            if "R1" in require_attention.attention_type:
                add_class_mark_into_order_line(order_line, "C2", "C", 1, 4)
            else:
                return
        else:
            return
    except Exception as e:
        logging.error(e)
        raise ValueError(INTERNAL_SERVER_ERROR)


def stamp_class_mark_for_dtr_dtp_logic(result_from_dtr_dtp, order_line):
    try:
        if result_from_dtr_dtp:
            add_class_mark_into_order_line(order_line, "C3", "C", 1, 4)
        else:
            return
    except Exception as e:
        logging.error(e)
        raise ValueError(INTERNAL_SERVER_ERROR)


def stamp_class_mark_for_change_order_domestic(request_date_is_change, order_line, flag):
    try:
        if request_date_is_change:
            add_class_mark_into_order_line(order_line, "C3" if flag == "Logistic" else "C4", "C", 1, 4)
        else:
            return
    except Exception as e:
        logging.error(e)
        raise ValueError(INTERNAL_SERVER_ERROR)


def stamp_class_mark_for_change_order_export(request_date_is_change, order_line):
    try:
        if request_date_is_change:
            add_class_mark_into_order_line(order_line, "C4", "C", 1, 4)
    except Exception as e:
        logging.error(e)
        raise ValueError(INTERNAL_SERVER_ERROR)


def stamp_class_mark_for_save_order_ref_contract(contract_name, order_line):
    try:
        if len(contract_name) < 2:
            return
        else:
            if contract_name[0:2] == 'DS' or contract_name[0:2] == 'DG' or contract_name[0:2] == 'DW':
                add_class_mark_into_order_line(order_line, "C1", "C", 1, 4)
            else:
                return
    except Exception as e:
        logging.error(e)
        raise ValueError(INTERNAL_SERVER_ERROR)


def stamp_class_mark_for_require_attention_change_request_date(is_request_date_change, order_line):
    if not is_request_date_change:
        return
    else:
        try:
            add_class_mark_into_order_line(order_line, "C4", "C", 1, 4)
        except Exception as e:
            logging.error(e)
            raise ValueError(INTERNAL_SERVER_ERROR)


def update_attention_type_r1(order_lines):
    """
    SEO-4434:
        Flag R1: request date < confirmed date or request date < confirm availability date
        Unflag R1: request >= confirmed date or request >= confirm availability date
    """
    for lines in order_lines:
        attention_type = lines.attention_type or ""
        attention_list = [att.strip() for att in attention_type.split(",") if att.strip()]
        r1_flagged = "R1" in attention_list
        if lines.request_date and lines.confirmed_date:
            request_before_confirmed = lines.request_date < lines.confirmed_date
            if request_before_confirmed and not r1_flagged:
                attention_list.append("R1")
            elif not request_before_confirmed and r1_flagged:
                attention_list.remove("R1")
        lines.attention_type = ", ".join(sorted(attention_list)).strip(", ")
    return order_lines


def update_attention_type_r3(order_lines):
    for lines in order_lines:
        attention_type = lines.attention_type
        if (
                lines.iplan
                and lines.sap_confirm_qty
                and lines.iplan.iplant_confirm_quantity
                and lines.iplan.on_hand_stock
        ):
            iplan_confirm_quantity = lines.iplan.iplant_confirm_quantity
            sap_confirm_qty = lines.sap_confirm_qty
            if iplan_confirm_quantity != sap_confirm_qty:
                attention_type = append_field(attention_type, "R3")
                attention_type = sort_values(attention_type, 1, 5, "R")
                lines.attention_type = attention_type
            if iplan_confirm_quantity == lines.sap_confirm_qty and validate_attention_type("R3", attention_type):
                tmp = lines.attention_type.split(", ")
                tmp.remove("R3")
                lines.attention_type = ", ".join(tmp)
    return order_lines


def parse_sap_date(date_str):
    for fmt in (DATE_FORMAT, '%Y%m%d'):
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            pass
    return None


def update_attention_type_r4(order_lines):
    for lines in order_lines:
        if lines.confirmed_date and lines.order.etd and lines.order.type == "export":
            attention_type = lines.attention_type
            confirmed_date = lines.confirmed_date
            etd = parse_sap_date(lines.order.etd)
            if not etd:
                continue
            if etd < confirmed_date:
                attention_type = append_field(attention_type, "R4")
                attention_type = sort_values(attention_type, 1, 5, "R")
                lines.attention_type = attention_type
            if etd >= confirmed_date and validate_attention_type("R4", attention_type):
                if attention_type.count(",") > 0:
                    new_attention_type = lines.attention_type.replace(", R4", "")
                else:
                    new_attention_type = lines.attention_type.replace("R4", "")
                lines.attention_type = new_attention_type
    return order_lines


def validate_attention_type(attention_type, field):
    if not field:
        return False
    return True if attention_type in field else False


def get_attention_type(order_line_attention_type, attention_type):
    if not order_line_attention_type:
        return attention_type
    if attention_type not in order_line_attention_type:
        return ', '.join(sorted(map(lambda x: x.strip(), f"{order_line_attention_type}, {attention_type}".split(","))))
    return order_line_attention_type


def update_attention_type_r5(order_lines):
    """
    Update type R5 for items fail confirm in i-plan
    @param order_lines: Line need update R5 attention
    """
    if not order_lines:
        # No order lines to update
        return
    update_lines = []
    order_id = None
    for line in order_lines:
        if not order_id:
            order_id = line.order.id
        line.attention_type = get_attention_type(line.attention_type, "R5")
        update_lines.append(line)
    sap_migrations_models.OrderLines.objects.bulk_update(update_lines, ["attention_type"])
    base_url = settings.MULESOFT_API_URL  # required
    iplan_split_url = base_url + IPlanEndPoint.I_PLAN_SPLIT.value
    iplan_confirm_url = base_url + IPlanEndPoint.I_PLAN_CONFIRM.value
    es21_url = base_url + SapEnpoint.ES_21.value
    filter_val = {
        "orderid": order_id,
        "url__in": (iplan_split_url, iplan_confirm_url, es21_url),
    }
    MulesoftApiRequest.instance().log_request_update(
        filter_val,  # Filter Values
        {"retry": True}  # Update Values
    )


def convert_to_ton(quantity, weight_unit, material_code):
    try:
        if weight_unit == 'TON':
            return quantity
        else:
            rate = master_data_models.Conversion2Master.objects.filter(from_unit=weight_unit,
                                                                       material_code=material_code,
                                                                       to_unit='TON').first()
        return quantity / rate.calculation
    except Exception:
        return quantity


def prepare_param_for_es21(so_no, order_lines, reason_for_reject):
    params = {
        "piMessageId": str(uuid.uuid1().int),
        "salesdocumentin": so_no,
        "testrun": "false",
        "orderHeaderIn": {
            "refDoc": "0410082672"

        },
        "orderHeaderInX": {},
        "orderItemsIn": [
            {
                "itemNo": order_line.get("item_no"),
                "material": order_line.get("material_variant_code"),
                "refDoc": "0410082672",
                "refDocIt": "000050",
            } for order_line in order_lines
        ]
    }
    for order_line in params.get('orderItemsIn'):
        if reason_for_reject == 'Delete':
            order_line['flgUpdateContract'] = 'D'
        elif reason_for_reject == 'Cancel 93':
            order_line['reasonReject'] = '93'

    return params


def is_valid_param_e21(param):
    return param.get("order__so_no") and param.get("item_no") and param.get("code")


def send_params_to_es21(key, value, manager, reason_for_reject):
    try:
        params = prepare_param_for_es21(key, value, reason_for_reject)
        MulesoftApiRequest.instance(service_type=MulesoftServiceType.SAP.value).request_mulesoft_post(
            SapEnpoint.ES_21.value,
            params
        )
    except Exception:
        pass


def is_valid_order_line_to_i_plan(order_line):
    return get_item_no_of_order_line(order_line) and get_product_code_of_order_line(order_line)


def get_params_by_inquiry_method(inquiry_method):
    if inquiry_method == InquiryMethodType.EXPORT.value:
        return {
            "inquiryMethod": IPlanInquiryMethodCode.JITCP.value,
            "useInventory": False,
            "useConsignmentInventory": True,
            "useProjectedInventory": True,
            "useProduction": True,
            "orderSplitLogic": "Split Multiple Date",
            "singleSourcing": True
        }
    if inquiry_method == InquiryMethodType.DOMESTIC.value or inquiry_method == InquiryMethodType.CUSTOMER.value:
        return {
            "inquiryMethod": IPlanInquiryMethodCode.JITCP.value,
            "useInventory": True,
            "useConsignmentInventory": True,
            "useProjectedInventory": True,
            "useProduction": True,
            "orderSplitLogic": "No Split",
            "singleSourcing": False
        }
    return {
        "inquiryMethod": IPlanInquiryMethodCode.ASAP.value,
        "useInventory": True,
        "useConsignmentInventory": True,
        "useProjectedInventory": True,
        "useProduction": True,
        "orderSplitLogic": "Split Multiple Date",
        "singleSourcing": True
    }


def get_inquiry_method_of_order_line(order_line):
    if order_line.inquiry_method:
        return order_line.inquiry_method
    try:
        code = order_line.order.distribution_channel.code
        if code in ["10", "20"]:
            return InquiryMethodType.DOMESTIC.value
        if code in ["30"]:
            return InquiryMethodType.EXPORT.value
    except AttributeError:
        return InquiryMethodType.ASAP.value


def get_product_code_of_order_line(order_line):
    variant = getattr(order_line, "material_variant", None)
    return getattr(variant, 'code', None)


def get_item_no_of_order_line(order_line):
    return getattr(order_line, 'item_no', None)


def prepare_for_ddq_request_line(order_line):
    template_ddq_request_line = {
        "lineNumber": get_item_no_of_order_line(order_line),
        "locationCode": "1012116",
        "productCode": get_product_code_of_order_line(order_line),
        "requestDate": "2022-10-28T03:20:08.873Z",
        "quantity": str(order_line.quantity),
        "unit": "ROL",
        "transportMethod": "",
        "typeOfDelivery": "A",
        "reATPRequired": False,
        "requestType": "NEW"
    }
    inquiry_method = get_inquiry_method_of_order_line(order_line)
    template_ddq_request_line.update(get_params_by_inquiry_method(inquiry_method))
    return template_ddq_request_line


def prepare_params_for_i_plan(list_order_line):
    if not list_order_line:
        return
    params = {
        "DDQRequest": {
            "requestId": str(uuid.uuid1().int),
            "sender": "e-ordering",
            "DDQRequestHeader": [
                {
                    "headerCode": str(uuid.uuid1().int),
                    "autoCreate": False,
                    "DDQRequestLine": [
                    ]
                }
            ]
        }
    }
    for order_line in list_order_line:
        if is_valid_order_line_to_i_plan(order_line):
            params["DDQRequest"]["DDQRequestHeader"][0]["DDQRequestLine"].append(
                prepare_for_ddq_request_line(order_line)
            )
    return params


def sale_org_list_param(sale_org_code, distribution_channel_code):
    value = {}
    if sale_org_code:
        value["salesOrg"] = sale_org_code
    if distribution_channel_code:
        value["distributionChannel"] = distribution_channel_code
    return value


def prepare_sale_org_list_for_es25(bu, sale_org, distribution_channel):
    result = []
    if sale_org == 'All':
        sale_org = master_data_models.SalesOrganizationMaster.objects.filter(business_unit__code=bu).values_list('code',
                                                                                                                 flat=True)
    if not isinstance(sale_org, QuerySet):
        sale_org = [sale_org]
    if distribution_channel == 'All':
        distribution_channel = ['10', '20']
    else:
        distribution_channel = [distribution_channel]
    for sale_org_code in sale_org:
        for distribution_channel_code in distribution_channel:
            result.append(sale_org_list_param(sale_org_code, distribution_channel_code))
    return result


def prepare_customer_list_for_es25(sold_to_code_input):
    result = []
    if sold_to_code_input:
        for sold_to_code in sold_to_code_input:
            sold_to_code = sold_to_code.split("-")[0].strip() if "-" in sold_to_code else sold_to_code
            result.append(sold_to_code)
    return result


def prepare_customer_list_for_api_es25(sold_to_code_input):
    result = []
    if sold_to_code_input:
        for sold_to_code in sold_to_code_input:
            if sold_to_code:
                param = {"customer": sold_to_code}
                result.append(param)
    return result


def prepare_params_for_field_can_choose_all_in_filter(field):
    if not field or field == 'All':
        return ""
    return field[1:] if len(field) == 4 else field


def prepare_material_for_es25(material_input):
    if not material_input:
        return ""
    else:
        material = master_data_models.MaterialMaster.objects.filter(material_code=material_input[0]).first()
        if material and material.material_type == '84':
            return f"{material_input[0]}*"
    return material_input[0]


def prepare_sale_employee_for_es25(sale_employee_input):
    result = ""
    if not sale_employee_input:
        return result
    for i in range(8):
        if sale_employee_input[i] == ' ':
            break
        result += sale_employee_input[i]
    return result


def prepare_material_group_1_for_es25(material_group_1_input):
    if not material_group_1_input or material_group_1_input == "All":
        return ""
    return material_group_1_input


def prepare_material_group_1_for_es25_order_pending(material_group_1_input):
    all_material_group = ProductGroup.get_all_product_groups().value
    if not material_group_1_input or set(material_group_1_input) == set(all_material_group):
        return ""
    return material_group_1_input[0]


def prepare_transactions_for_es25(transaction_input):
    block = "Delivery block"
    normal = "Normal"
    all_transactions = [block, normal]
    if not transaction_input or set(transaction_input) == set(all_transactions):
        return ""
    if transaction_input[0] == block:
        return "Block"
    return transaction_input[0]


def prepare_material_pricing_group_for_es25(material_pricing_group_group_input):
    mapping = {
        "Standard": "01",
        "Non standard": "02"
    }
    return mapping.get(material_pricing_group_group_input, "")


def prepare_customer_po_no_for_es25(customer_po_no_input):
    if not customer_po_no_input:
        return ""
    return customer_po_no_input


def convert_date_to_fit_params_in_es25(string_date):
    a = string_date.split("-")
    return '/'.join(a[::-1])


def add_key_and_data_into_params_in_es25(key, value, params):
    if not value:
        return
    params[key] = value


def make_list_sale_org(bu):
    return master_data_models.SalesOrganizationMaster.objects.filter(business_unit__code=bu).values_list("code",
                                                                                                         flat=True)


def fill_zeros_prepare_param(filter_input):
    if not filter_input:
        return ""
    return filter_input.zfill(10)


def prepare_param_for_es25(filter_input, get_order_line=False):
    params = {
        "piMessageId": str(uuid.uuid1().int),
    }
    bu = filter_input.get('bu', "")
    if not bu:
        raise ValueError("bu is mandatory")
    list_sale_org = make_list_sale_org(bu)
    # Prepare parameters
    sales_order_no = fill_zeros_prepare_param(filter_input.get('sale_order_no', ""))
    sale_org_and_channel = prepare_sale_org_list_for_es25(bu, filter_input.get('sale_org', list_sale_org),
                                                          filter_input.get('channel', ""))
    customer = prepare_customer_list_for_api_es25(filter_input.get('sold_to', []))
    sales_employee = prepare_sale_employee_for_es25(filter_input.get('sales_employee_no'))
    sales_group = prepare_params_for_field_can_choose_all_in_filter(filter_input.get('sale_group', ""))
    pending_status = filter_input.get('status', "")
    customer_po_no = filter_input.get('purchase_order_no', "")
    create_date_from = convert_date_to_fit_params_in_es25(str(filter_input.get('create_date', {}).get('gte', "")))
    create_date_to = convert_date_to_fit_params_in_es25(str(filter_input.get('create_date', {}).get('lte', "")))
    request_date_from = convert_date_to_fit_params_in_es25(
        str(filter_input.get('request_delivery_date', {}).get('gte', "")))
    request_date_to = convert_date_to_fit_params_in_es25(
        str(filter_input.get('request_delivery_date', {}).get('lte', "")))
    order_type = filter_input.get('order_type', "")
    sales_order_type = "" if order_type == 'All' else order_type
    material = prepare_material_for_es25(filter_input.get('material_no_material_description', ""))
    material_group_1 = prepare_params_for_field_can_choose_all_in_filter(filter_input.get('material_group_1', ""))
    plant = filter_input.get('plant', "")
    mat_pricing_grp = prepare_material_pricing_group_for_es25(filter_input.get('material_pricing_group', ""))
    contract_no = fill_zeros_prepare_param(filter_input.get('contract_no', ""))
    dp_no = fill_zeros_prepare_param(filter_input.get('dp_no', ""))
    last_update_date_from = convert_date_to_fit_params_in_es25(
        str(filter_input.get('last_update_date', {}).get('gte', "")))
    last_update_date_to = convert_date_to_fit_params_in_es25(
        str(filter_input.get('last_update_date', {}).get('lte', "")))
    ship_to = filter_input.get('ship_to', "")
    billing_no = fill_zeros_prepare_param(filter_input.get('invoice_no', ""))
    created_by = prepare_created_by_for_es25(filter_input.get('create_by', None))

    # Add key-value to params
    add_key_and_data_into_params_in_es25("salesOrderNo", sales_order_no, params)
    add_key_and_data_into_params_in_es25("salesOrgList", sale_org_and_channel, params)
    add_key_and_data_into_params_in_es25("customerList", customer, params)
    add_key_and_data_into_params_in_es25("salesEmployee", sales_employee, params)
    add_key_and_data_into_params_in_es25("salesGroup", sales_group, params)
    add_key_and_data_into_params_in_es25("pendingStatus", pending_status, params)
    add_key_and_data_into_params_in_es25("customerPoNo", customer_po_no, params)
    add_key_and_data_into_params_in_es25("createDateFrom", create_date_from, params)
    add_key_and_data_into_params_in_es25("createDateTo", create_date_to, params)
    add_key_and_data_into_params_in_es25("requestDateFrom", request_date_from, params)
    add_key_and_data_into_params_in_es25("requestDateTo", request_date_to, params)
    add_key_and_data_into_params_in_es25("salesOrderType", sales_order_type, params)
    add_key_and_data_into_params_in_es25("material", material, params)
    add_key_and_data_into_params_in_es25("materialGroup1", material_group_1, params)
    add_key_and_data_into_params_in_es25("plant", plant, params)
    add_key_and_data_into_params_in_es25("matPricingGrp", mat_pricing_grp, params)
    add_key_and_data_into_params_in_es25("contractNo", contract_no, params)
    add_key_and_data_into_params_in_es25("dpNo", dp_no, params)
    add_key_and_data_into_params_in_es25("lastUpdateDateFrom", last_update_date_from, params)
    add_key_and_data_into_params_in_es25("lastUpdateDateTo", last_update_date_to, params)
    add_key_and_data_into_params_in_es25("shipTo", ship_to, params)
    add_key_and_data_into_params_in_es25("billingNo", billing_no, params)
    add_key_and_data_into_params_in_es25("webUserId", created_by, params)
    source_of_app = get_filter_source_of_app(filter_input)
    if source_of_app:
        add_key_and_data_into_params_in_es25("sourceApplication", source_of_app, params)
    if get_order_line:
        add_key_and_data_into_params_in_es25("flgItem", "X", params)

    return params


def get_filter_source_of_app(filter_input):
    input_source_app = filter_input.get('source_of_app', "")
    if (input_source_app and input_source_app.upper() == "ALL"):
        return None
    else:
        return input_source_app


def get_request_delivery_date_of_order_line_in_database(order_line):
    if not order_line:
        return ""
    return order_line.order.request_delivery_date if order_line.order else ""


def get_original_request_date_of_order_line_in_database(order_line):
    return getattr(order_line, 'original_request_date', "")


def get_overdue_1_of_order_line_in_database(order_line):
    return check_overdue_required_attention(order_line, "request_date")


def get_overdue_2_of_order_line_in_database(order_line):
    return check_overdue_required_attention(order_line, "confirmed_date")


def check_overdue_required_attention(order_line, date_field):
    date_field_data = getattr(order_line, date_field, None)
    if order_line and order_line.attention_type and date_field_data:
        current_date = datetime.now().date()
        is_contain_r1 = "R1" in order_line.attention_type
        if current_date > date_field_data and is_contain_r1:
            return True
    return False


def get_required_attention_flag_of_order_line_in_database(order_line):
    return getattr(order_line, 'attention_type', "")


def get_created_by_of_order_line_in_database(order_line):
    if not order_line:
        return ""
    created_by = order_line.order.created_by
    if created_by:
        return f"{created_by.first_name} {created_by.last_name}"
    return ""


def get_confirm_date_of_order_line_in_database(order_line):
    if not order_line:
        return ""
    return order_line.confirmed_date.strftime(DATE_FORMAT_ISO) if order_line.confirmed_date else ""


def get_sale_group_of_order_line_in_database(order_line):
    order = getattr(order_line, "order", None)
    sales_group = getattr(order, "sales_group", None)
    return getattr(sales_group, "code", "")


def get_request_date_of_order_line_in_database(order_line):
    return getattr(order_line, "request_date", '')


def get_company_name():
    name_list = sap_master_data.models.SalesOrganizationMaster.objects.all().values("code", "short_name")
    mapping_data = {sale_org['code']: sale_org for sale_org in name_list}
    return mapping_data


def mapping_material_group1(order_lines, product_group_filter):
    if product_group_filter:
        return list(filter(lambda x: x.get('materialGroup1', "") in product_group_filter, order_lines))
    # CIP order items
    for item in order_lines:
        item["materialGroup1"] = "All Product"
    return order_lines


def mapping_product_group(list_order_line):
    result = {}
    for line in list_order_line:
        product_group = line["materialGroup1"]
        try:
            item = ProductGroupDescription[product_group].value if product_group in ProductGroupDescription.__members__ \
                else product_group
        except KeyError:
            item = None
        if item is None:
            continue
        result.setdefault(item, {'product_group': product_group, 'order_lines': []}).get('order_lines').append(line)
    return result


def create_date_time_for_list_sales_order(date, time):
    return str(date) + ' - ' + str(time)


def check_line_is_container(line: dict):
    if line.get("itemCategory") == ItemCat.ZKC0.value:
        return True
    return False


def make_order_line_for_list_order_line(order_line, mapping_order_with_order_no):
    sd_doc = order_line.get("sdDoc")
    item_no = order_line.get("itemNo")
    original_request_date = order_line.get("shiptToPODate", "")

    order_line_in_database = sap_migrations_models.OrderLines.objects.filter(order__so_no=sd_doc,
                                                                             item_no=item_no.lstrip("0")).first()
    sale_org_code = mapping_order_with_order_no.get(sd_doc).get('salesOrg', "")
    e_ordering_confirm_date = getattr(order_line_in_database, "confirmed_date", "")

    if is_bom_parent(order_line.get("bomFlag", ""), order_line.get("parentItemNo", "")):
        child_order_line = OrderLineRepo.get_order_line_by_order_and_parent_item_no(sd_doc, item_no.lstrip("0"))
        e_ordering_confirm_date = getattr(child_order_line, "confirmed_date", "")

    if e_ordering_confirm_date:
        e_ordering_confirm_date = e_ordering_confirm_date.strftime(DATE_FORMAT_ISO)  # noqa
    sale_org_short_name = sap_master_data.models.SalesOrganizationMaster.objects.filter(
        code=sale_org_code).first().short_name
    status = order_line.get('status', "")
    net_value = order_line.get("netValue", 0)
    if isinstance(net_value, str):
        net_value = float(net_value)

    order = mapping_order_with_order_no.get(sd_doc, {})
    return {
        "create_date_time": order_line.get('createDate', None),
        "sales_org": f"{sale_org_code} - {sale_org_short_name}",
        "sales_group": order.get("salesGroup", ""),
        "sold_to": f"{order.get('soldTo')} - {order.get('soldToName1')}",
        "po_no": order.get('poNo'),
        "original_request_date": original_request_date,  # Original Date
        "create_date": order.get('createDate'),
        "req_delivery_date": order_line.get("firstDeliveryDT", ""),  # วันที่ต้องการ สินค้า
        "order_no": sd_doc,
        "item_no": item_no.lstrip("0"),
        "mat_no": order_line.get("matNo", ""),
        "price_date": order_line.get("priceDate", ""),
        "bom_flag": BOM_FLAG_TRUE_VALUE == order_line.get("bomFlag", ""),
        "parent_item_no": order_line.get("parentItemNo", "").lstrip("0"),
        "mat_desc": order_line.get("matDesc", ""),
        "material_pricing_group": order_line.get("materialPricingGroup", ""),
        "order_qty": scg_round(order_line.get("orderQty", 0)),
        "confirm_order_qty": order_line.get('confirmQty') and scg_round(order_line.get("confirmQty")),
        "delivery_qty": order_line.get('giQty') and scg_round(order_line.get('giQty')),
        "pending_qty": order_line.get('pendingGiQty') and scg_round(order_line.get("pendingGiQty")),
        "unit": order_line.get("saleUnit", ""),
        "plant": order_line.get("plant", ""),
        "shipping_point": order_line.get("shippingPoint", ""),
        "order_weight": order_line.get("netWeightTon", 0),
        "delivery_weight": order_line.get("giQtyTon", 0),
        "pending_weight": order_line.get("pendingGiQtyTon", 0),
        "weight_unit": order_line.get("weightUnitTon", ""),
        "net_price": f'{order_line.get("netPrice", 0):.2f}',
        "net_value": f'{net_value:.2f}',
        "currency": order_line.get("priceCurrency", ""),
        "rejection": order_line.get("reasonReject", ""),
        "delivery_block": "Block" if order_line.get("deliveryBlock", "") == "09" else "Unblock",
        "e_ordering_overdue_1": get_overdue_1_of_order_line_in_database(order_line_in_database),
        "e_ordering_overdue_2": get_overdue_2_of_order_line_in_database(order_line_in_database),
        "e_ordering_required_attention_flag": get_required_attention_flag_of_order_line_in_database(
            order_line_in_database),
        "e_ordering_create_by": order.get('webUserId'),
        "e_ordering_confirm_date": e_ordering_confirm_date,  # วันที่ประมาณการส่งมอบ
        'confirm_qty': scg_round(order_line.get("orderQty", 0)),
        'ship_to_name': order.get('shipToName1'),
        'e_ordering_request_date': getattr(order_line_in_database, "confirmed_date", ""),
        "e_ordering_item_status": getattr(order_line_in_database, "item_status_en", ''),
        "iplan_confirm_date": getattr(order_line_in_database, "confirmed_date", ""),
        "sap_status": SAP_STATUS_MAPPING.get(status, ""),
        "is_not_ref": order.get('is_not_ref', False),
        "order_tracking_status": order_line.get("trackingStatus", ""),
        "tracking_url": order_line.get("URL", "")
    }


def calculate_qty_base(order_line, unit, qty_base):
    if check_line_is_container(order_line):
        return "-"
    return f'{(float(qty_base) / 1000 if unit == "KG" else qty_base):.3f}'


def make_list_order_line_for_sap_order_line(list_order_line_of_sold_to, mapping_order_with_order_no):
    res = []
    for order_line in list_order_line_of_sold_to:
        res.append(make_order_line_for_list_order_line(order_line, mapping_order_with_order_no))
    return res


def make_summary_unit(order_lines, select_unit):
    order_lines_filter = list(filter(lambda x: x.get(select_unit), order_lines))
    units = map(lambda line: line.get(select_unit), order_lines_filter)
    return "/".join(sorted(list(set(units))))


def make_summary_for_sold_to_from_order_line_in_es25(list_order_line):
    order_weight = delivery_weight = pending_weight = net_price = net_value = 0
    set_currency = set()
    for order_line in list_order_line:
        if is_bom_child(order_line.get("bomFlag", ""), order_line.get("parentItemNo", "")):
            continue
        if order_line.get("priceCurrency"):
            set_currency.add(order_line.get("priceCurrency"))
        net_price += float(order_line.get("netPrice", 0))
        net_value += float(order_line.get("netValue", 0))
        order_weight += 0 if (check_line_is_container(order_line)) else float(
            order_line.get("netWeightTon", 0))
        delivery_weight += 0 if (check_line_is_container(order_line)) else float(
            order_line.get("giQtyTon", 0))
        pending_weight += 0 if (check_line_is_container(order_line)) else float(
            order_line.get("pendingGiQtyTon", 0))
    currency = ''
    for item in set_currency:
        currency += item + '/'

    # verification required new field
    weight_unit = make_summary_unit(list_order_line, "weightUnitTon")
    quantity_data = quantity_data_calculation_from_order_lines(list_order_line)

    return {
        'quantity_data': quantity_data,
        'order_weight': f'{order_weight:.3f}',
        'delivery_weight': f'{delivery_weight:.3f}',
        'pending_weight': f'{pending_weight:.3f}',
        'weight_unit': weight_unit,
        'net_price': f'{net_price:.2f}',
        'net_value': f'{net_value:.2f}',
        'currency': currency[:-1],
    }


def quantity_data_calculation_from_order_lines(list_order_line):
    data = []
    unit_data = {}
    try:
        for order_line in list_order_line:
            if is_bom_child(order_line.get("bomFlag", ""), order_line.get("parentItemNo", "")):
                continue
            unit = order_line.get("saleUnit")
            if unit not in unit_data:
                unit_data[unit] = {
                    'confirm_order_qty': 0,
                    'order_qty': 0,
                    'delivery_qty': 0,
                    'pending_qty': 0,
                }
            unit_data[unit]['confirm_order_qty'] += float(
                order_line.get("confirmQty") if order_line.get("confirmQty") else 0)
            unit_data[unit]['order_qty'] += float(order_line.get("orderQty") if order_line.get("orderQty") else 0)
            unit_data[unit]['delivery_qty'] += float(order_line.get("giQty") if order_line.get("giQty") else 0)
            unit_data[unit]['pending_qty'] += float(
                order_line.get("pendingGiQty") if order_line.get("pendingGiQty") else 0)
        for unit, values in unit_data.items():
            data.append({
                'confirm_order_qty': f'{values["confirm_order_qty"]:{CURRENCY_FORMAT_CONSTANT}}',
                'order_qty': f'{values["order_qty"]:{CURRENCY_FORMAT_CONSTANT}}',
                'delivery_qty': f'{values["delivery_qty"]:{CURRENCY_FORMAT_CONSTANT}}',
                'pending_qty': f'{values["pending_qty"]:{CURRENCY_FORMAT_CONSTANT}}',
                'quantity_unit': unit,
            })
    except Exception as e:
        logging.error(e)
    return data


def get_list_partner_from_order(partner_code_list):
    from scgp_export.implementations.orders import get_name_address_from_sold_to_partner

    list_partner = {}
    try:
        list_sold_to_partner_address = (
            master_data_models.SoldToPartnerAddressMaster.objects.filter(
                partner_code__in=partner_code_list,
            )
        ).distinct("partner_code").in_bulk(field_name="partner_code")
        for partner_code, sold_to_partner_address in list_sold_to_partner_address.items():
            list_partner[partner_code] = get_name_address_from_sold_to_partner(sold_to_partner_address)

        return list_partner

    except Exception:
        return {}


def mapping_consignment_list(summary_response):
    res = []

    dummy_list_res = summary_response.get("DummyListResponseLine", [])
    for record in dummy_list_res:
        res.append({
            "customer_session": f'{record.get("customerCode", "")} - {record.get("customerName", "")}',
            "quantity": scg_round(record.get("onHandQuantity", 0)),
            "future_dummy_stock": scg_round(record.get("futureQuantity", 0)),
            "plant": f'{record.get("warehouseCode", "")} - {PLANT_MAPPING.get(record.get("warehouseCode", "")).get("name2", "")}',
            # No data response now
        })
    return res


def mapping_free_stock_list(summary_response):
    res = []

    free_stock_list_res = summary_response.get("FreeStockResponseLine", [])
    logging.info(f"Free stock response: {free_stock_list_res}")
    for record in free_stock_list_res:
        # SEO-6370 prod issue
        warehouse_code = record.get("warehouseCode", "")
        if PLANT_MAPPING.get(warehouse_code):
            plant = f'{warehouse_code} - {PLANT_MAPPING.get(warehouse_code).get("name2", "")}'
        else:
            plant = f'{warehouse_code}'
        res.append({
            "plant": plant,
            "quantity": scg_round(record.get("onHandQuantity", 0)),
        })
    return res


def mapping_ipan_65201_response(api_response, code):
    summary_response = api_response.get("OnHandCSInquiryResponse", {}).get("OnHandCSInquirySummaryResponse", {})
    customer_order_quantity = summary_response.get("customerOrdersQuantity")
    dummy_quantity = summary_response.get("dummyQuantity")
    free_quantity = summary_response.get("freeQuantity")
    total_ur = customer_order_quantity + dummy_quantity + free_quantity
    unit = summary_response.get("unit")

    # All data are mocked, need to update when API is ready
    res = {
        "product_code": code,
        "total_ur": f"{float(total_ur):.3f}",
        "customer_order_quantity": scg_round(customer_order_quantity),
        "consignment_quantity": scg_round(dummy_quantity),
        "free_quantity": scg_round(free_quantity),
        "unit": unit,
        "summary_as_of": django_tz.now().strftime(DATE_FORMAT_ISO),  # Mock, no reponse
        "customer_order_list": mapping_customer_order_list(summary_response),
        "consignment_list": mapping_consignment_list(summary_response),
        "free_stock_list": mapping_free_stock_list(summary_response),
    }
    return res


def prepare_sale_org_list_for_es25_order_pending(filter_input):
    """
        CS Domestics and Customer [Not include CS Export]
            CS Domestics and Customer with distributionChannel is 10 or 20
            CS Export with distributionChannel is 30
    """
    sale_org = filter_input.get('sale_org')
    if sale_org and "All" in sale_org:
        bu = filter_input.get('bu')
        sale_org = master_data_models.SalesOrganizationMaster.objects.filter(business_unit__code=bu).values_list('code',
                                                                                                                 flat=True)
    res = []
    for code in sale_org:
        res.extend([
            {"salesOrg": code, "distributionChannel": "10"},
            {"salesOrg": code, "distributionChannel": "20"}
        ])
    return res


def prepare_param_for_es25_order_pending(filter_input):
    source_of_app = get_filter_source_of_app(filter_input)
    params = {
        "piMessageId": str(uuid.uuid1().int),
        "salesOrderNo": filter_input.get('so_no', ""),
        "shipTo": filter_input.get('ship_to', ""),
        "salesOrgList": prepare_sale_org_list_for_es25_order_pending(filter_input),
        "customerList": [{"customer": customer} for customer in
                         prepare_customer_list_for_es25(filter_input.get('sold_to'))],
        "pendingStatus": "G",
        "customerPoNo": prepare_customer_po_no_for_es25(filter_input.get('po_no')),
        "createDateFrom": convert_date_to_fit_params_in_es25(str(filter_input.get('create_date', {}).get('gte', ""))),
        "createDateTo": convert_date_to_fit_params_in_es25(str(filter_input.get('create_date', {}).get('lte', ""))),
        "material": prepare_material_for_es25(filter_input.get('material_no_material_description')),
        "materialGroup1": prepare_material_group_1_for_es25(filter_input.get('product_groups')),
        "deliveryBlockStatus": prepare_transactions_for_es25(filter_input.get('transactions')),
        "requestDateFrom": convert_date_to_fit_params_in_es25(
            str(filter_input.get('request_delivery_date', {}).get('gte', ""))),
        "requestDateTo": convert_date_to_fit_params_in_es25(
            str(filter_input.get('request_delivery_date', {}).get('lte', ""))),
        "flgItem": "X",
        "sourceApplication": source_of_app
    }
    params = {k: v for k, v in params.items() if v}
    return params


def scg_round(value):
    value = float(value)
    if value == int(value):
        return int(value)
    value = "%.3f" % value
    return int(float(value)) if float(value).is_integer() else value


def scg_round_2f(value):
    value = float(value)
    if value == int(value):
        return int(value)
    value = "%.2f" % value
    return int(float(value)) if float(value).is_integer() else value


def make_order_line_for_list_order_line_pending(order_line, orders):
    company_name = get_company_name()
    sale_unit = order_line.get("saleUnit", "")
    order = next((x for x in orders if x.get("sdDoc") == order_line.get("sdDoc")), {})
    sales_org = order.get('salesOrg', {})
    sales_org_str = str(sales_org)
    return {
        "sale_org": f"{sales_org_str} - {company_name.get(sales_org_str, {}).get('short_name', '')}",
        "sold_to": f"{order.get('soldTo')} - {order.get('soldToName1')}",
        "ship_to": f"{order.get('shipTo')} - {order.get('shipToName1')}",
        "request_date": order_line.get("shiptToPODate", ""),
        "order_date": order.get("createDate", ""),
        "po_no": order.get("poNo", ""),
        "so_no": order_line.get("sdDoc", ""),
        "item_no": order_line.get("itemNo", ""),
        "material_code": order_line.get("matNo", ""),
        "material_description": order_line.get("matDesc", ""),
        "order_qty": scg_round(order_line.get("orderQty", 0)),
        "pending_qty": order_line.get("pendingGiQty") and scg_round(order_line.get("pendingGiQty")),
        "atp_qty": order_line.get("confirmQty") and scg_round(order_line.get("confirmQty")),
        "ctp_qty": order_line.get("nonConfirmQty") and scg_round(order_line.get("nonConfirmQty")),
        "sale_unit": sale_unit,
        "ship_to_code": order.get('shipTo', ""),
        "ship_to_name": order.get('shipToName1', ""),
        "confirm_date": order_line.get("firstDeliveryDT", ""),
        "delivery_qty": order_line.get("giQty") and scg_round(order_line.get("giQty")),
        "ship_to_po_date": order_line.get("shiptToPODate", ""),
        "bom_flag": BOM_FLAG_TRUE_VALUE == order_line.get("bomFlag"),
        "parent_item_no": order_line.get("parentItemNo", "")
    }


def mapping_customer_order_list(summary_response):
    res = []
    so_no_from_res = []
    customer_list = set()
    customer_orders_res = summary_response.get("CustomerOrdersResponseLine", [])

    # Collect customer codes
    for record in customer_orders_res:
        so_no_from_res.append(record.get("orderNumber", "").zfill(10))
        customer_code = record.get("customerCode", "")
        if customer_code:
            customer_list.add(customer_code.zfill(10))

    mapped_orders = {order.so_no: order for order in
                     sap_migrations_models.Order.objects.filter(so_no__in=so_no_from_res)}
    # Get customer names and addresses
    list_customers_name_and_address = get_list_partner_from_order(customer_list)

    # Map customer order details
    for record in customer_orders_res:
        so_no = record.get("orderNumber", "").zfill(10)
        order = mapped_orders.get(so_no)
        request_date = ""
        delivery_date = record.get("deliveryDate")
        if delivery_date:
            request_date = datetime.strptime(delivery_date, DATE_FORMAT).strftime(DATE_FORMAT_ISO)

        customer_code = record.get("customerCode", "").zfill(10)
        customer_name = list_customers_name_and_address.get(customer_code, {}).get("name")

        res_dict = {
            "customer": f'{customer_code} - {customer_name}',
            "so_no_slash_item": f'{so_no} / {record.get("lineNumber", "")}',
            "quantity": scg_round(record.get("reATPQuantity", 0)),
            "pending_quantity": scg_round(record.get("pendingDeliveryQuantity", 0)),
            "request_date": request_date,
            "plant": f'{record.get("warehouseCode", "")} - {PLANT_MAPPING.get(record.get("warehouseCode", "")).get("name2", "")}',
            "so_no": so_no,
            "create_date": "",
            "create_by": None,
            "sale_rep": "",
            "po_no": "",
        }
        if order:
            res_dict["create_date"] = order.created_at.strftime(DATE_FORMAT_ISO) if order.created_at else ""
            res_dict["create_by"] = order.created_by
            res_dict["sale_rep"] = order.sales_employee
            if order.type in {"domestic", "customer"}:
                res_dict["po_no"] = order.po_number
            else:
                res_dict["po_no"] = order.po_no
        res.append(res_dict)
    return res


def make_summary_for_sold_to_from_order_line_in_es25_order_pending(list_order_line):
    summary = []
    sale_unit_groups = {}
    # group items by sale_unit
    for order_line in list_order_line:
        sale_unit = order_line['saleUnit']
        sale_unit_groups.setdefault(sale_unit, []).append(order_line)
    # prepare summary for each sale_unt
    for sale_unit, order_lines in sale_unit_groups.items():
        order_qty = delivery_qty = pending_qty = atp_qty = ctp_qty = 0
        for order_line in order_lines:
            # exclude bom child from summary calculation
            if is_bom_child(order_line.get("bomFlag"), order_line.get("parentItemNo")):
                continue
            order_qty += float(order_line.get("orderQty") if order_line.get("orderQty") else 0)
            pending_qty += float(order_line.get("pendingGiQty") if order_line.get("pendingGiQty") else 0)
            delivery_qty += float(order_line.get("giQty") if order_line.get("giQty") else 0)
            atp_qty += float(order_line.get("confirmQty") if order_line.get("confirmQty") else 0)
            ctp_qty += float(order_line.get("nonConfirmQty") if order_line.get("nonConfirmQty") else 0)
        summary.append(
            {
                'order_qty': scg_round(order_qty),
                'pending_qty': scg_round(pending_qty),
                'atp_qty': scg_round(atp_qty),
                'ctp_qty': scg_round(ctp_qty),
                'delivery_qty': scg_round(delivery_qty),
                'sale_unit': sale_unit,
            }
        )
    return summary


def make_list_order_line_for_sap_order_line_pending(order_lines, orders, sort_columns, sold_to):
    res = []
    default_sort_field_list = ['so_no', 'item_no']
    for order_line in order_lines:
        res.append(make_order_line_for_list_order_line_pending(order_line, orders))
    if sort_columns:
        bom_child_list, parent_and_normal_order_line_list = separate_parent_and_bom_order_lines(res)
        sorted_parent_and_normal_order_line_list = sort_pending_order_lines(default_sort_field_list,
                                                                            parent_and_normal_order_line_list, sold_to,
                                                                            sort_columns)
        sorted_bom_child_list = sort_pending_order_lines(default_sort_field_list, bom_child_list, sold_to, sort_columns)
        reverse = sort_columns[0].get("sort_type") == "DESC"
        res = sorted_and_merged_order_line_list(reverse, sorted_bom_child_list,
                                                sorted_parent_and_normal_order_line_list)

    else:
        res = sorted(res, key=lambda x: multiple_sort_for_pending_order(x, default_sort_field_list))

    return res


def sort_pending_order_lines(default_sort_field_list, res, sold_to, sort_columns):
    for sort_column in sort_columns:
        if sort_column.get("sold_to") == '' or sold_to == sort_column.get("sold_to"):
            column = PendingOrderFieldHeaderColumn().find("en", sort_column.get("sort_field"))
            reverse = sort_column.get("sort_type") == "DESC"
            if column:
                if camel_to_snake(column['en']) in default_sort_field_list:
                    default_sort_field_list.remove(camel_to_snake(column['en']))
                default_sort_field_list.insert(0, camel_to_snake(column['en']))
            res = sorted(res,
                         key=lambda x: multiple_sort_for_pending_order(x, default_sort_field_list, reverse=reverse),
                         reverse=reverse)
    return res


def group_otc_orders(orders):
    otc_order_groups = {}
    for otc_order in orders:
        otc_order_groups.setdefault(otc_order.get("soldToName1"), []).append(otc_order)
    return otc_order_groups


def group_items_by_sold_to(orders, data_item, sort_columns):
    orders_no = list(map(lambda x: x.get("sdDoc"), orders))
    order_lines = list(filter(lambda x: x.get("sdDoc") in orders_no, data_item))
    order = next(iter(orders), {})
    row = {
        "sold_to": order.get("soldTo"),
        "sold_to_name": order.get("soldToName1", ""),
        "product_groups": []
    }
    groups = mapping_product_group(order_lines)
    for product_group_desc, data in groups.items():
        order_line_data = data['order_lines']
        if order_line_data:
            inner_data = {
                "product_group": product_group_desc,
                "material_group": data['product_group'],
                "order_lines": make_list_order_line_for_sap_order_line_pending(order_line_data, orders,
                                                                               sort_columns,
                                                                               order.get("soldTo")),
                "summary": make_summary_for_sold_to_from_order_line_in_es25_order_pending(order_line_data)
            }
            row['product_groups'].append(inner_data)
    return row


def remove_parent_item_no_for_child_search(order_lines):
    items = list(map(lambda x: x.get("sdDoc") + x.get("itemNo"), order_lines))
    for order_line in order_lines:
        if order_line.get("parentItemNo") and order_line.get("sdDoc") + order_line.get("parentItemNo") not in items:
            del order_line["parentItemNo"]
            del order_line["bomFlag"]


def multiple_sort_for_pending_order(item, sort_list, is_index=False, reverse=False):
    sort_parameters = []
    secondary_sorting_fields = ['so_no', 'item_no']
    if sort_list:
        for i, column in enumerate(sort_list):
            field = PendingOrderFieldHeaderColumn().find("en", snake_to_camel(column))
            col_type = field['type']
            if is_index:
                index = field['index']
                value = item[index]
            else:
                value = item[column]
            if col_type == "date":
                if value:
                    field_value = _convert_to_datetime(value)
                else:
                    if reverse:
                        # If sorting in reverse, None should be smallest (will come last)
                        field_value = datetime.min
                    else:
                        # If sorting normally, None should be largest (will come last)
                        field_value = datetime.max
            elif col_type == "int":
                field_value = int(value)
            elif col_type == "float":
                field_value = float(value)
            else:
                field_value = str(value).lower()
            # ascending order secondary-third sorting fields always SEO-7427
            if reverse and i > 0 and column in secondary_sorting_fields:
                field_value = int(field_value) * -1
            sort_parameters.append(field_value)
    return tuple(sort_parameters)


def _convert_to_datetime(date_str):
    return datetime.strptime(date_str, "%d/%m/%Y")


def camel_to_snake(camel_case):
    snake_case = re.sub(r'(?<!^)(?=[A-Z])', '_', camel_case).lower()
    return snake_case


def format_date_for_excel(input_date):
    try:
        date_obj = datetime.strptime(input_date, DATE_FORMAT_ISO) if isinstance(input_date, str) else input_date
        return date_obj.strftime("%d.%m.%Y")
    except Exception:
        return ""


def file_headers_for_sold_to_group_in_list_sales_order(is_order_tracking=False):
    tracking_status = ("Order Tracking Status",) if is_order_tracking else ()
    price_date = ("Pricing date",) if is_order_tracking else ()

    return [
        "Required attention flag",
        "Item Status",
        "SAP Status",
        *tracking_status,
        "Create By",
        "Report create date-time",
        "Sale org",
        "Sales Group",
        "Sold to",
        "PO No.",
        "Original Request Date",
        "Create Date",
        "Req.Delivery Date",
        "Confirm date",
        "Order No.",
        "Order Item",
        "Material Code",
        "Material Description",
        "Order Qty.",
        "Confirm Order Qty.",
        "Delivery QTY.",
        "Pending Qty.",
        "Unit",
        "Plant",
        "Shipping Point",
        "Order Weight",
        "Delivery Weight",
        "Pending Weight",
        "Weight Unit",
        "Net Price",
        "Net Value",
        "Currency",
        *price_date,
        "Rejection",
        "Delivery block",
        "Overdue 1",
        "Overdue 2"
    ]


def file_headers_for_summary_in_list_sales_order():
    return [
        'Order Qty.',
        'Confirm Order Qty.',
        'Delivery QTY.',
        'Pending  Qty.',
        'Unit',
        'Order Weight',
        'Delivery Weight',
        'Pending Weight',
        'Weight Unit',
        'Net Price ',
        'Net Value',
        'Currency'
    ]


def map_data_for_each_row_in_list_sales_order(order_line, is_order_tracking=False):
    if is_order_tracking:
        tracking_status = (order_line.get("order_tracking_status", ""),)
        order_line["price_date"] = order_line.get("price_date").replace("/", ".") if order_line.get(
            "price_date") and order_line.get("price_date") != "" else ""
        price_date = (order_line.get("price_date", ""),)
    else:
        tracking_status = ()
        price_date = ()

    return [
        order_line["e_ordering_required_attention_flag"],
        order_line["e_ordering_item_status"],
        order_line["sap_status"],
        *tracking_status,
        order_line["e_ordering_create_by"],
        datetime.now(timezone("Asia/Bangkok")).strftime("%d.%m.%Y/%H:%M:%S"),
        order_line["sales_org"],
        order_line["sales_group"],
        order_line["sold_to"],
        order_line["po_no"],
        format_date_for_excel(order_line["original_request_date"]),
        format_date_for_excel(order_line["create_date"]),
        format_date_for_excel(order_line["req_delivery_date"]),
        format_date_for_excel(order_line["e_ordering_confirm_date"]),
        order_line["order_no"],
        order_line["item_no"],
        order_line["mat_no"],
        order_line["mat_desc"],
        order_line["order_qty"],
        order_line["confirm_order_qty"],
        order_line["delivery_qty"],
        order_line["pending_qty"],
        order_line["unit"],
        order_line["plant"],
        order_line["shipping_point"],
        order_line["order_weight"],
        order_line["delivery_weight"],
        order_line["pending_weight"],
        order_line["weight_unit"],
        order_line["net_price"],
        order_line["net_value"],
        order_line["currency"],
        *price_date,
        order_line["rejection"],
        order_line["delivery_block"],
        "Yes" if order_line["e_ordering_overdue_1"] else "",
        "Yes" if order_line["e_ordering_overdue_2"] else "",
    ]


def make_excel_from_list_of_sale_order_sap(list_of_sale_order_sap, is_order_tracking):
    file_headers = file_headers_for_sold_to_group_in_list_sales_order(is_order_tracking)
    temporary_file = create_file_with_headers(file_headers, ",", "xlsx")
    for group in list_of_sale_order_sap:
        row_group = [["Sold to - Code - Name"]]
        order_lines = group['order_lines']
        sold_to_code_name = order_lines[0]["sold_to"]
        row_group[0].append("")
        row_group[0].append("")
        if is_order_tracking:
            row_group[0].append("")
        row_group[0].append(sold_to_code_name)
        for order_line in order_lines:
            _row = map_data_for_each_row_in_list_sales_order(order_line, is_order_tracking)
            row_group.append(_row)
        row_summary = ["Summary"]
        for i in range(16 if is_order_tracking else 15):
            row_summary.append("")
        summary = group['summary']
        _row_val_for_summary = []
        i = 0
        for key, val in summary.items():
            i += 1
            if i == 2:
                row_summary.append("")
                row_summary.append("Total (ทั้งหมด)")
            if key == 'quantity_data':
                data = val[0]
                if data:
                    row_summary.append("Total(" + data['quantity_unit'] + ")")
                    row_summary.append(data['order_qty'])
                    row_summary.append(data['confirm_order_qty'])
                    row_summary.append(data['delivery_qty'])
                    row_summary.append(data['pending_qty'])
                    row_summary.append(data['quantity_unit'])
                else:
                    row_summary.append("")
                    row_summary.append("")
                    row_summary.append("")
                    row_summary.append("")
                    row_summary.append("")
                    row_summary.append("")
            else:
                row_summary.append(val)
        row_group.append(row_summary)

        quantity_data = summary['quantity_data'][1:]
        for data in quantity_data:
            row_summary = []
            i = 0
            for i in range(17 if is_order_tracking else 16):
                row_summary.append("")
            row_summary.append("Total(" + data['quantity_unit'] + ")")
            row_summary.append(data['order_qty'])
            row_summary.append(data['confirm_order_qty'])
            row_summary.append(data['delivery_qty'])
            row_summary.append(data['pending_qty'])
            row_summary.append(data['quantity_unit'])
            row_group.append(row_summary)

        etl.io.xlsx.appendxlsx(row_group, temporary_file.name, write_header=True)

    set_excel_column_style_for_sales_order_report(temporary_file)

    with open(temporary_file.name, "rb") as exported_file:
        base64_file = base64.b64encode(exported_file.read())
    return base64_file.decode("utf-8")


def set_excel_column_style_for_sales_order_report(temporary_file):
    wb = openpyxl.load_workbook(temporary_file)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(values_only=False)):
        row_bold = False
        for cell in row:
            if isinstance(cell.value, str):
                if "Required attention flag" in cell.value:
                    set_style(ws, f'{cell.column_letter}', cell.row, alignment_style=True, auto_size=False)
                    ws.row_dimensions[cell.row].height = 50
                if "Item Status" in cell.value:
                    set_style(ws, 'B', cell.row, 27, False)
                if "SAP Status" in cell.value:
                    set_style(ws, f'{cell.column_letter}', cell.row, width=16, alignment_style=False, auto_size=True)
                if "Order Tracking Status" in cell.value:
                    set_style(ws, f'{cell.column_letter}', cell.row, width=14, alignment_style=True, auto_size=False)
                if "Create By" in cell.value:
                    set_style(ws, column=f'{cell.column_letter}', row_number=cell.row, width=54, alignment_style=False)
                if "Report create date-time" in cell.value:
                    set_style(ws, column=f'{cell.column_letter}', row_number=cell.row, width=24, alignment_style=False)
                if "Sale org" in cell.value:
                    set_style(ws, column=f'{cell.column_letter}', row_number=cell.row, width=27, alignment_style=False)
                if "Sales Group" in cell.value:
                    set_style(ws, column=f'{cell.column_letter}', row_number=cell.row, width=8, alignment_style=True,
                              auto_size=False)
                if "Sold to" == cell.value:
                    set_style(ws, column=f'{cell.column_letter}', row_number=cell.row, width=82, alignment_style=False)
                if "PO No." in cell.value:
                    set_style(ws, column=f'{cell.column_letter}', row_number=cell.row, width=23, alignment_style=True,
                              auto_size=False,
                              wrap_cell_value=True,
                              min_row=cell.row + 1, min_col=cell.column, max_col=cell.column,
                              cell_value_alignment_horizontal_left=True)
                if "Original Request Date" in cell.value:
                    set_style(ws, column=f'{cell.column_letter}', row_number=cell.row, alignment_style=True,
                              auto_size=False)
                if "Create Date" in cell.value:
                    set_style(ws, column=f'{cell.column_letter}', row_number=cell.row, alignment_style=False,
                              auto_size=True)
                if "Req.Delivery Date" in cell.value:
                    set_style(ws, column=f'{cell.column_letter}', row_number=cell.row, alignment_style=True,
                              auto_size=False)
                if "Confirm date" in cell.value:
                    set_style(ws, column=f'{cell.column_letter}', row_number=cell.row, alignment_style=False,
                              auto_size=True)
                if "Order No." in cell.value:
                    set_style(ws, column=f'{cell.column_letter}', row_number=cell.row, width=13, alignment_style=False)
                if "Order Item" in cell.value:
                    set_style(ws, column=f'{cell.column_letter}', row_number=cell.row, alignment_style=False,
                              auto_size=True)
                if "Material Code" in cell.value:
                    set_style(ws, column=f'{cell.column_letter}', row_number=cell.row, width=27, alignment_style=False)
                if "Material Description" in cell.value:
                    set_style(ws, column=f'{cell.column_letter}', row_number=cell.row, width=54, alignment_style=False)
                if "Order Qty." in cell.value:
                    set_style(ws, column=f'{cell.column_letter}', row_number=cell.row, width=13, alignment_style=False,
                              number_format='#,##0.000',
                              two_decimal=False, min_row=2, min_col=cell.column, max_col=cell.column)
                if "Confirm Order Qty." in cell.value:
                    set_style(ws, column=f'{cell.column_letter}', row_number=cell.row, width=13, alignment_style=True,
                              auto_size=False,
                              number_format='#,##0.000', min_row=2, min_col=cell.column, max_col=cell.column,
                              two_decimal=False)
                if "Delivery QTY." in cell.value:
                    set_style(ws, column=f'{cell.column_letter}', row_number=cell.row, width=13, alignment_style=False,
                              number_format='#,##0.000', min_row=2, min_col=cell.column, max_col=cell.column,
                              two_decimal=False)
                if "Pending Qty." in cell.value:
                    set_style(ws, column=f'{cell.column_letter}', row_number=cell.row, width=13, alignment_style=False,
                              number_format='#,##0.000', min_row=2, min_col=cell.column, max_col=cell.column,
                              two_decimal=False)
                if "Unit" in cell.value:
                    set_style(ws, column=f'{cell.column_letter}', row_number=cell.row, width=10, alignment_style=False)
                if "Plant" in cell.value:
                    set_style(ws, column=f'{cell.column_letter}', row_number=cell.row, width=10, alignment_style=False)
                if "Shipping Point" in cell.value:
                    logging.info(f"cell value {cell.value} and cell letter {cell.column_letter}")
                    set_style(ws, column=f'{cell.column_letter}', row_number=cell.row, width=12, alignment_style=True,
                              auto_size=False,
                              cell_value_alignment_horizontal_left=True, min_row=2, min_col=cell.column,
                              max_col=cell.column)
                if "Order Weight" in cell.value:
                    logging.info(f"cell value {cell.value} and cell letter {cell.column_letter}")
                    set_style(ws, column=f'{cell.column_letter}', row_number=cell.row, width=10, alignment_style=True,
                              number_format='#,##0.000', min_row=2, min_col=cell.column, max_col=cell.column,
                              two_decimal=False)
                if "Delivery Weight" in cell.value:
                    logging.info(f"cell value {cell.value} and cell letter {cell.column_letter}")
                    set_style(ws, column=f'{cell.column_letter}', row_number=cell.row, width=10, alignment_style=True,
                              number_format='#,##0.000', min_row=2, min_col=cell.column, max_col=cell.column,
                              two_decimal=False)
                if "Pending Weight" in cell.value:
                    logging.info(f"cell value {cell.value} and cell letter {cell.column_letter}")
                    set_style(ws, column=f'{cell.column_letter}', row_number=cell.row, width=10, alignment_style=True,
                              number_format='#,##0.000', min_row=2, min_col=cell.column, max_col=cell.column,
                              two_decimal=False)
                if "Weight Unit" in cell.value:
                    logging.info(f"cell value {cell.value} and cell letter {cell.column_letter}")
                    set_style(ws, column=f'{cell.column_letter}', row_number=cell.row, width=10, alignment_style=True,
                              auto_size=False)
                if "Net Price" in cell.value:
                    logging.info(f"cell value {cell.value} and cell letter {cell.column_letter}")
                    set_style(ws, column=f'{cell.column_letter}', row_number=cell.row, width=15, alignment_style=False,
                              number_format='#,##0.00', min_row=2, min_col=29, max_col=29, two_decimal=True,
                              alignment_horizontal_right=True)
                if "Net Value" in cell.value:
                    logging.info(f"cell value {cell.value} and cell letter {cell.column_letter}")
                    set_style(ws, column=f'{cell.column_letter}', row_number=cell.row, width=15, alignment_style=False,
                              number_format='#,##0.00', min_row=2, min_col=30, max_col=30, two_decimal=True,
                              alignment_horizontal_right=True)
                if "Currency" in cell.value:
                    logging.info(f"cell value {cell.value} and cell letter {cell.column_letter}")
                    set_style(ws, column=f'{cell.column_letter}', row_number=cell.row, alignment_style=True,
                              auto_size=' False')
                if "Pricing date" in cell.value:
                    set_style(ws, column=f'{cell.column_letter}', row_number=cell.row, alignment_style=True,
                              auto_size=' False')

                if "Rejection" in cell.value:
                    set_style(ws, column=f'{cell.column_letter}', row_number=cell.row, alignment_style=True,
                              auto_size=' False',
                              cell_value_alignment_horizontal_left=True, min_row=1, min_col=cell.column,
                              max_col=cell.column)
                if "Delivery block" in cell.value:
                    set_style(ws, column=f'{cell.column_letter}', row_number=cell.row, width=10, alignment_style=True,
                              auto_size=' False')
                if "Overdue 1" in cell.value:
                    set_style(ws, column=f'{cell.column_letter}', row_number=cell.row, alignment_style=True,
                              auto_size=' False',
                              cell_value_alignment_horizontal_left=True, min_row=1, min_col=cell.column,
                              max_col=cell.column)
                if "Overdue 2" in cell.value:
                    set_style(ws, column=f'{cell.column_letter}', row_number=cell.row, alignment_style=True,
                              auto_size=' False',
                              cell_value_alignment_horizontal_left=True, min_row=1, min_col=cell.column,
                              max_col=cell.column)
    wb.save(temporary_file.name)


def get_order_from_order_line(order_line):
    return getattr(order_line, 'order', None)


def set_style(ws, column, row_number, width=None, alignment_style=False, auto_size=None, wrap_cell_value=None,
              min_row=None, min_col=None, max_col=None, number_format=None, two_decimal=None,
              alignment_horizontal_right=None, height=None, cell_value_alignment_horizontal_left=None, ):
    ws[f'{column}{row_number}'].alignment = ALIGNMENT_STYLE if alignment_style else ALIGNMENT_VERTICAL_TOP
    if height:
        ws.row_dimensions[row_number].height = height
    if width:
        ws.column_dimensions[column].width = width

    if number_format:
        for col in ws.iter_cols(min_row=min_row, min_col=min_col, max_col=max_col):
            for cell in col:
                if cell.value and two_decimal is not None:
                    try:
                        cell.value = scg_round_2f(cell.value) if two_decimal else scg_round(cell.value)
                        if cell.value and min_col in [29, 30]:
                            cell.value = float(cell.value)
                            cell.number_format = number_format
                            ws[f'{column}{cell.row}'].alignment = Alignment(horizontal='left', vertical='top')
                    except ValueError:
                        print("Unable to convert cell value to float.")

                if isinstance(cell.value, int) and min_col in [8, 9, 10, 11]:
                    ws[f'{column}{cell.row}'].alignment = Alignment(wrap_text=True, horizontal='right', vertical='top')

                if type(cell.value) is not None and min_col in [25, 26, 27]:
                    cell.number_format = number_format
                    ws[f'{column}{cell.row}'].alignment = Alignment(horizontal='right', vertical='top')
                    try:
                        if cell.value and not isinstance(cell.value, int):
                            cell.value = float(cell.value)
                    except ValueError:
                        print("Unable to convert cell value to float.")
                    cell.number_format = number_format
                    if "Total (ทั้งหมด)" == ws[f'X{cell.row}'].value:
                        cell.number_format = number_format
                        ws[f'{column}{cell.row}'].alignment = Alignment(horizontal='left', vertical='top')

                if type(cell.value) is not None and min_col in [18, 19, 20, 21]:
                    if cell.value and isinstance(cell.value, str):
                        try:
                            cell.value = float(cell.value)
                        except ValueError:
                            print("Unable to convert cell value to float.")
                        cell.number_format = number_format
                        ws[f'{column}{cell.row}'].alignment = Alignment(horizontal='left', vertical='top')
                    if "Total(ROL)" == ws[f'Q{cell.row}'].value:
                        cell.number_format = number_format
                        ws[f'{column}{cell.row}'].alignment = Alignment(horizontal='left', vertical='top')
                if cell.value and isinstance(cell.value, str) and min_col in [8, 9, 10, 11]:
                    try:
                        cell.value = float(cell.value)
                    except ValueError:
                        print("Unable to convert cell value to float.")
                    cell.number_format = number_format
                    ws[f'{column}{cell.row}'].alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')

    if cell_value_alignment_horizontal_left:
        for col in ws.iter_cols(min_row=min_row, min_col=min_col, max_col=max_col):
            for cell in col:
                ws[f'{column}{cell.row}'].alignment = Alignment(horizontal='left', vertical='top')

    if wrap_cell_value:
        for row_sub in ws.iter_rows(min_row=min_row, min_col=min_col, max_col=max_col):
            for cell_sub in row_sub:
                cell_sub.alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal='left', vertical='top')

    if auto_size is not None:
        ws.column_dimensions[column].auto_size = auto_size


def get_material_variant_from_order_line(order_line):
    return getattr(order_line, 'material_variant', None)


def get_material_from_order_line(order_line):
    material_variant = get_material_variant_from_order_line(order_line)
    return getattr(material_variant, 'material', None)


def get_sold_to_from_order_line(order_line):
    order = get_order_from_order_line(order_line)
    contract = getattr(order, 'contract', None)
    return getattr(contract, "sold_to", None)


def get_sales_organization_from_order_line(order_line):
    order = get_sold_to_from_order_line(order_line)
    return getattr(order, "sales_organization", None)


def get_sales_group_from_order_line(order_line):
    order = get_order_from_order_line(order_line)
    return getattr(order, 'sales_group', None)


def get_i_plan_from_order_line(order_line):
    return getattr(order_line, "iplan", None)


def get_order_line_from_order_line_iplan(order_line_iplan):
    return getattr(order_line_iplan, "orderlines", None)


def prepare_sale_org_list_for_es25_order_confirmation(filter_input):
    sale_org = filter_input.get('sale_org')
    distribution_channel = filter_input.get('channel')
    if sale_org and "All" in sale_org:
        bu = filter_input.get('bu')
        sale_org = master_data_models.SalesOrganizationMaster.objects.filter(business_unit__code=bu).values_list('code',
                                                                                                                 flat=True)
    res = []
    for code in sale_org:
        res.append(
            {
                "salesOrg": code,
                "distributionChannel": distribution_channel
            }
        )
    return res


def get_distribution_channel_id_from_code(code):
    if not code:
        return None
    instance = master_data_models.DistributionChannelMaster.objects.filter(code=code).first()
    return instance or None


def get_sales_org_id_from_code(code):
    if not code:
        return None
    instance = master_data_models.SalesOrganizationMaster.objects.filter(code=code).first()
    return instance or None


def get_division_id_from_code(code):
    if not code:
        return None
    instance = master_data_models.DivisionMaster.objects.filter(code=code).first()
    return instance or None


def get_distribution_channels_by_codes(codes):
    if not codes:
        return None
    instance = master_data_models.DistributionChannelMaster.objects.filter(code__in=codes).all()
    return instance or None


def get_sales_orgs_by_codes(codes):
    if not codes:
        return None
    instance = master_data_models.SalesOrganizationMaster.objects.filter(code__in=codes).all()
    return instance or None


def get_incoterms_id_from_code(code):
    if not code:
        return None
    instance = master_data_models.Incoterms1Master.objects.filter(code=code).first()
    return instance or None


def get_customer_group_id_from_code(code):
    if not code:
        return None
    instance = master_data_models.CustomerGroupMaster.objects.filter(code=code).first()
    return instance or None


def get_customer_group1_id_from_code(code):
    if not code:
        return None
    instance = master_data_models.CustomerGroup1Master.objects.filter(code=code).first()
    return instance or None


def get_sales_group_id_from_code(code):
    if not code:
        return None
    instance = sap_migrations_models.SalesGroupMaster.objects.filter(code=code).first()
    return instance or None


def get_sales_office_group_id_from_code(code):
    if not code:
        return None
    instance = sap_migrations_models.SalesOfficeMaster.objects.filter(code=code).first()
    return instance or None


def get_currency_id_from_code(code):
    if not code:
        return None
    instance = sap_migrations_models.CurrencyMaster.objects.filter(code=code).first()
    return instance or None


def get_contract_id_from_contract_no(contract_no):
    if not contract_no:
        return None
    instance = sap_migrations_models.Contract.objects.filter(code=contract_no).first()
    return instance or None


def get_sold_to_id_from_contract_no(contract_no):
    if not contract_no:
        return None
    instance = sap_migrations_models.Contract.objects.filter(code=contract_no).first()
    return getattr(instance, "sold_to", None)


def sync_from_es26_to_database(data_from_es26):
    date_format_es26 = DATE_FORMAT_ISO
    date_format_to_database = DATE_FORMAT
    _order_from_es26 = data_from_es26["data"][0]["orderHeaderIn"]
    if not _order_from_es26:
        return
    _order_partners_from_es26 = data_from_es26["data"][0]["orderPartners"]
    _order_lines_from_es26 = data_from_es26["data"][0]["orderItems"]
    list_material_code = []
    list_contract_material_code = []
    for order_line in _order_lines_from_es26:
        material_code = order_line.get("material")
        contract_code = order_line.get("contractNo")
        if material_code:
            list_material_code.append(material_code)
        if contract_code:
            list_contract_material_code.append(contract_code)

    mapping_material_variant_code_with_material_variant_id = sap_migrations_models.MaterialVariantMaster.objects.filter(
        code__in=list_material_code).distinct("code").order_by("code").in_bulk(field_name="code")

    mapping_contract_material_code_with_contract_material = sap_migrations_models.ContractMaterial.objects.filter(
        contract_no__in=list_contract_material_code).distinct(
        "contract_no").order_by("contract_no").in_bulk(field_name="contract_no")
    order_in_database = sap_migrations_models.Order.objects.filter(so_no=_order_from_es26["saleDocument"]).first()
    if not order_in_database:
        order_in_database = sap_migrations_models.Order.objects.create(
            so_no=_order_from_es26["saleDocument"],
            distribution_channel=get_distribution_channel_id_from_code(_order_from_es26["distributionChannel"]),
            sales_organization=get_sales_org_id_from_code(_order_from_es26["salesOrg"]),
            division=get_division_id_from_code(_order_from_es26["division"]),
            request_date=datetime.strptime(_order_from_es26["reqDate"], date_format_es26).strftime(
                date_format_to_database),
            incoterms_1=get_incoterms_id_from_code(_order_from_es26["incoterms1"]),
            incoterms_2=_order_from_es26["incoterms2"],
            payment_term=_order_from_es26["paymentTerms"],
            po_no=_order_from_es26["poNo"],
            price_group=_order_from_es26["priceGroup"],
            price_date=datetime.strptime(_order_from_es26["priceDate"], date_format_es26).strftime(
                date_format_to_database),
            created_at=datetime.strptime(_order_from_es26["createDate"], date_format_es26).strftime(
                date_format_to_database),
            customer_group=get_customer_group_id_from_code(_order_from_es26["customerGroup"]),
            customer_group_1=get_customer_group1_id_from_code(_order_from_es26.get("customerGroup1")),
            sales_district=_order_from_es26["salesDistrict"],
            shipping_condition=_order_from_es26["shippingCondition"],
            sales_group=get_sales_group_id_from_code(_order_from_es26["salesGroup"]),
            sales_office=get_sales_office_group_id_from_code(_order_from_es26["salesOff"]),
            currency=get_currency_id_from_code(_order_from_es26["currency"]),
            contract=get_contract_id_from_contract_no(_order_from_es26.get("contractNo")),
            sold_to=get_sold_to_id_from_contract_no(_order_from_es26.get("contractNo"))
        )
        order_in_database.save()
    list_update_fields = [
        "item_no",
        "order",
        "material_variant",
        "quantity",
        "sales_unit",
        "plant",
        "payment_term_item",
        "po_no",
        "item_category",
        "request_date",
        "shipping_point",
        "reject_reason",
        "contract_material"
    ]
    list_order_line_create = []
    list_order_line_update = []
    list_item_no = []
    for order_line in _order_lines_from_es26:
        list_item_no.append(order_line["itemNo"].lstrip("0"))
    item_no_of_order_line_exits_in_database = sap_migrations_models.OrderLines.objects.filter(
        order=order_in_database, item_no__in=list_item_no).distinct(
        "item_no").order_by("item_no").in_bulk(field_name="item_no")
    for order_line in _order_lines_from_es26:
        _order_line = sap_migrations_models.OrderLines(
            item_no=order_line["itemNo"].lstrip("0"),
            order=order_in_database,
            material_variant=mapping_material_variant_code_with_material_variant_id.get(order_line.get("material")),
            quantity=order_line.get("itemQty", None),
            sales_unit=order_line.get("salesUnit", None),
            plant=order_line.get("plant", None),
            payment_term_item=order_line.get("paymentTerm", None),
            po_no=order_line.get("poNumber", None),
            item_category=order_line.get("itemCategory", None),
            request_date=datetime.strptime(order_line.get("requestedDate"), date_format_es26).strftime(
                date_format_to_database),
            shipping_point=order_line.get("shippingPoint", None),
            reject_reason=order_line.get("reasonReject", None),
            contract_material=mapping_contract_material_code_with_contract_material.get(
                order_line.get("contractNo")),
        )
        if order_line["itemNo"].lstrip("0") not in item_no_of_order_line_exits_in_database:
            list_order_line_create.append(_order_line)
        if order_line["itemNo"].lstrip("0") in item_no_of_order_line_exits_in_database:
            item_no = order_line["itemNo"].lstrip("0")
            _order_line.id = item_no_of_order_line_exits_in_database.get(item_no).id
            list_order_line_update.append(_order_line)
            del item_no_of_order_line_exits_in_database[order_line["itemNo"].lstrip("0")]
    sap_migrations_models.OrderLines.objects.filter(
        order=order_in_database).exclude(item_no__in=list_item_no).delete()
    if list_order_line_create:
        sap_migrations_models.OrderLines.objects.bulk_create(list_order_line_create)
    if list_order_line_update:
        sap_migrations_models.OrderLines.objects.bulk_update(list_order_line_update, list_update_fields)


def update_remark_order_line(order_line_remark, remark):
    if not order_line_remark:
        return remark
    if remark not in order_line_remark:
        return ', '.join(sorted(map(lambda x: x.strip(), f"{order_line_remark}, {remark}".split(","))))
    return order_line_remark


def dtr_dtp_stamp_class_mark(order_line_ids, class_mark):
    order_lines = sap_migrations_models.OrderLines.objects.filter(id__in=order_line_ids).all()
    for order_line in order_lines:
        order_line.remark = update_remark_order_line(order_line.remark, class_mark)
    sap_migrations_models.OrderLines.objects.bulk_update(order_lines, fields=["remark"])


def prepare_created_by_for_es25(user_id):
    user = User.objects.filter(id=user_id).first()
    if user:
        return f'{user.first_name} {user.last_name}'
    return ""


def format_item_no(item):
    item['itemNo'] = item['itemNo'].lstrip("0")
    if item.get('parentItemNo'):
        item['parentItemNo'] = item['parentItemNo'].lstrip("0")
    return item


def derive_parent_price_and_weight_for_bom(current_order_lines):
    parent_item_no_child_item_dict = {}
    for current_order_line in current_order_lines:
        parent_item_no = current_order_line.get("parentItemNo", "")
        if is_bom_child(current_order_line.get("bomFlag", ""), parent_item_no):
            parent_item_no_child_item_dict.setdefault(parent_item_no, []).append(current_order_line)
    if parent_item_no_child_item_dict:
        for current_order_line in current_order_lines:
            if is_bom_parent(current_order_line.get("bomFlag", ""), current_order_line.get("parentItemNo", "")):
                child_lines = parent_item_no_child_item_dict.get(current_order_line.get("itemNo"), "")
                if child_lines:
                    parent_order_weight = parent_delivery_weight = parent_pending_weight = parent_net_price = parent_net_value = 0
                    for child_line in child_lines:
                        parent_order_weight += child_line.get("netWeightTon", 0)
                        parent_delivery_weight += child_line.get("giQtyTon", 0)
                        parent_pending_weight += child_line.get("pendingGiQtyTon", 0)
                        parent_net_value += child_line.get("netValue", 0)
                        parent_net_price += child_line.get("netPrice", 0)
                        child_line['netValue'] = 0
                        child_line['netPrice'] = 0
                    current_order_line['netWeightTon'] = parent_order_weight
                    current_order_line['giQtyTon'] = parent_delivery_weight
                    current_order_line['pendingGiQtyTon'] = parent_pending_weight
                    current_order_line['netValue'] = parent_net_value
                    current_order_line['netPrice'] = parent_net_value / scg_round(current_order_line.get("orderQty", 1))


def set_default_quantity_to_es25_response(es25_response_items):
    for item in es25_response_items:
        item["giQty"] = item.get("giQty", 0)
        item["pendingGiQty"] = item.get("pendingGiQty", 0)
        item["confirmQty"] = item.get("confirmQty", 0)
        item["nonConfirmQty"] = item.get("nonConfirmQty", 0)
